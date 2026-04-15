from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

DARK = "1B4332"
MED = "2D6A4F"
LIGHT = "D8F3DC"
WHITE = "FFFFFF"
GRAY = "F5F5F5"
BLUE = "E3F2FD"
YELLOW = "FFF9C4"
GREEN = "E8F5E9"
ORANGE = "FFF3E0"
RED_BG = "FFEBEE"
BORDER = "BDBDBD"

hdr_font = Font(name="Arial", bold=True, color=WHITE, size=11)
hdr_fill = PatternFill("solid", fgColor=DARK)
sub_font = Font(name="Arial", bold=True, color=DARK, size=10)
sub_fill = PatternFill("solid", fgColor=LIGHT)
body = Font(name="Arial", size=10)
bold = Font(name="Arial", size=10, bold=True)
code_font = Font(name="Consolas", size=9, color="1B5E20")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="top", wrap_text=True)
border = Border(
    left=Side("thin", BORDER), right=Side("thin", BORDER),
    top=Side("thin", BORDER), bottom=Side("thin", BORDER)
)

def hdr(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.alignment = center; cell.border = border

def row_style(ws, row, cols, fill=None):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = body; cell.alignment = wrap; cell.border = border
        if fill: cell.fill = PatternFill("solid", fgColor=fill)

def section_row(ws, row, cols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = Font(name="Arial", bold=True, color=WHITE, size=11)
    cell.fill = PatternFill("solid", fgColor=MED)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border = border
    ws.row_dimensions[row].height = 26

# ════════════════════════════════════════
# SHEET 1: ALL PAGES
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "Pages (60)"
ws1.sheet_properties.tabColor = DARK

cols1 = ["#", "Page Path", "Module", "Lines", "Type", "Key Features", "State Variables", "API Calls", "Components Used"]
for i, h in enumerate(cols1, 1): ws1.cell(row=1, column=i, value=h)
hdr(ws1, 1, len(cols1))

pages = [
    # Dashboard
    [1, "/", "Dashboard", "~500", "Page", "Admin dashboard with stats cards, today's appointments, recent patients, revenue summary", "stats, appointments, patients, loading", "/api/dashboard", "StatsCard, LayoutShell"],
    # Patients
    [2, "/patients", "Patients", "410", "List", "Patient list with search, status/gender filters, sorting, table+cards", "patients, search, statusFilter, genderFilter, sortField, sortDir", "/api/patients?search=&status=", "FilterChip, SortHeader"],
    [3, "/patients/new", "Patients", "~1400", "Form", "30+ field registration, NRIC checksum validation, duplicate detection (NRIC/phone/email), family members, medical history, SG address, DOB/age toggle", "saving, dobMode, selectedConditions, selectedGroups, duplicates, duplicateOverrides, fieldErrors, familyMembers, isDirty", "/api/patients (POST), /api/patients/check-duplicate, /api/patients?search=", "DuplicateWarning (inline)"],
    [4, "/patients/[id]", "Patients", "~3881", "Detail", "18-tab patient profile: appointments, clinical notes, treatment plans, vitals, documents, billing, timeline, family, photo, edit mode", "patient, activeSection (18 tabs), editing, editData, clinicalNotes, documents, vitals, familyMembers, timeline, invoices, payments", "/api/patients/[id], /api/patients/[id]/timeline, /api/patients/[id]/family, /api/clinical-notes, /api/documents, /api/vitals, /api/invoices", "ConfirmDialog"],
    # Appointments
    [5, "/appointments", "Appointments", "~600", "List", "Appointment calendar, list view, status filters, doctor filter", "appointments, view (calendar/list), statusFilter, doctorFilter", "/api/appointments, /api/doctors", ""],
    [6, "/appointments/new", "Appointments", "~500", "Form", "Book appointment with doctor selection, slot picker, walk-in support", "doctors, selectedDoctor, date, availableSlots, selectedSlot", "/api/doctors, /api/doctors/[id]/slots, /api/appointments (POST)", ""],
    # Billing
    [7, "/billing", "Billing", "~800", "List", "Invoice list with branch filter, status filter, search, stats cards, sorting", "invoices, stats, branchFilter, statusFilter, search, sortField", "/api/invoices, /api/billing/stats, /api/branches", "BillingTabs"],
    [8, "/billing/new", "Billing", "~1200", "Form", "Create invoice: patient search, treatment/medicine items with variants, GST calc, discounts, package linkage", "patient, items, treatments, medicines, variants, totals", "/api/patients, /api/treatments, /api/inventory/lookup, /api/invoices (POST)", ""],
    [9, "/billing/[id]", "Billing", "~1000", "Detail", "Invoice view: items, payments, credit notes, edit, print, payment recording", "invoice, payments, creditNotes, editing", "/api/invoices/[id], /api/invoices/[id]/payments, /api/credit-notes", "BillingTabs, ConfirmDialog"],
    [10, "/billing/insurance", "Billing", "~400", "List", "Insurance claims list with status filter", "claims, statusFilter", "/api/insurance/claims", "BillingTabs"],
    [11, "/billing/insurance/providers", "Billing", "~400", "List", "Insurance provider management", "providers", "/api/insurance/providers", "BillingTabs"],
    # Inventory
    [12, "/inventory", "Inventory", "594", "List", "Inventory list with search, category filter, stock status badges, stats cards, A-Z sorting", "items, search, categoryFilter, statusFilter, stats, sortField", "/api/inventory, /api/inventory/stats", "InventoryTabs"],
    [13, "/inventory/new", "Inventory", "393", "Form", "Create item: name, category, subcategory (30+ for medicine), unit, pricing, stock, GST, expiry, supplier", "form fields, categories, subcategories, suppliers", "/api/inventory (POST), /api/suppliers", "InventoryTabs"],
    [14, "/inventory/[id]", "Inventory", "675+", "Detail", "Item detail with 4 sub-tabs: Overview, Variants, Transactions, Movement Graph. Edit capability", "item, activeTab, transactions, variants, movementData, editing", "/api/inventory/[id], /api/inventory/[id]/transactions, /api/inventory/[id]/movement", "InventoryTabs"],
    [15, "/inventory/stock-audit", "Inventory", "395+", "Form", "Physical stock count with barcode scanning, branch selection, discrepancy detection", "auditItems, branch, physicalCounts, results", "/api/inventory/stock-audit (POST), /api/branches", "InventoryTabs, BarcodeScanner"],
    [16, "/inventory/alerts", "Inventory", "400+", "Dashboard", "Low stock + expiry alerts, write-off actions, PO creation from alerts", "alerts (lowStock, expired, expiringSoon, outOfStock), branchFilter", "/api/inventory/alerts, /api/inventory/write-off-expired", "InventoryTabs"],
    [17, "/inventory/branch-stock", "Inventory", "591", "Dashboard", "Branch stock comparison, imbalance detection, transfer suggestions, export", "comparisonData, branches, categoryFilter, showDiffsOnly", "/api/inventory/branch-comparison, /api/branches", "InventoryTabs"],
    [18, "/inventory/import", "Inventory", "400+", "Form", "CSV/Excel import with column mapping, validation, preview, batch insert", "file, mappings, previewData, validationErrors, importResults", "/api/inventory/import (POST)", "InventoryTabs"],
    [19, "/inventory/reports", "Inventory", "400+", "Dashboard", "4 report tabs: Stock Summary, Low Stock, Expiry Analysis, Transfers", "activeTab, reportData, dateRange, branchFilter", "/api/reports/inventory, /api/reports/transfers", "InventoryTabs"],
    [20, "/inventory/suppliers", "Inventory", "400", "List/CRUD", "Supplier directory with add/edit modal, status toggle, order stats", "suppliers, showModal, form, editingId", "/api/suppliers (GET/POST/PUT)", "InventoryTabs"],
    [21, "/inventory/purchase-orders", "Inventory", "400+", "List", "PO list with status filter, search, stats, smart suggestions link", "purchaseOrders, statusFilter, stats", "/api/purchase-orders", "InventoryTabs"],
    [22, "/inventory/purchase-orders/new", "Inventory", "597", "Form", "Create PO: supplier, branch, items with autocomplete, quantities, prices, GST, totals", "supplier, branch, lineItems, totals, savingAs (draft/submitted)", "/api/purchase-orders (POST), /api/suppliers, /api/inventory/lookup, /api/branches, /api/purchase-orders/suggestions", "InventoryTabs"],
    [23, "/inventory/purchase-orders/[id]", "Inventory", "400+", "Detail", "PO detail: items, status workflow, receiving with barcode scan, notes", "po, receiving, receivedItems", "/api/purchase-orders/[id], /api/purchase-orders/[id]/receive, /api/purchase-orders/[id]/status", "InventoryTabs, BarcodeScanner"],
    [24, "/inventory/transfers", "Inventory", "435", "List", "Transfer list with status/date filters, template dropdown, stats", "transfers, statusFilter, dateRange, stats", "/api/transfers, /api/transfers/templates", "InventoryTabs"],
    [25, "/inventory/transfers/new", "Inventory", "586", "Form", "Create transfer: branch selection, stock lookup, item search, qty validation (≤ available)", "fromBranch, toBranch, branchStock, lineItems, notes", "/api/transfers (POST), /api/branches, /api/branches/stock", "InventoryTabs"],
    [26, "/inventory/transfers/[id]", "Inventory", "400+", "Detail", "Transfer detail: items, status actions (submit/receive/cancel), receiving with barcode", "transfer, receiving, receivedItems", "/api/transfers/[id], /api/transfers/[id]/submit|receive|cancel", "InventoryTabs, BarcodeScanner"],
    # Doctor Portal
    [27, "/doctor", "Doctor Portal", "351", "Dashboard", "Doctor dashboard: stats, today's appointments, upcoming (7d), recent prescriptions, low stock alerts", "dashboardData, lowStockItems, activeTab", "/api/doctor/dashboard, /api/branches/stock, /api/inventory", ""],
    [28, "/doctors", "Doctors", "~400", "List", "Doctor directory with specialization filter", "doctors, search", "/api/doctors", ""],
    [29, "/doctors/new", "Doctors", "~300", "Form", "Add doctor (now via User model with role=doctor)", "form fields", "/api/staff (POST)", ""],
    [30, "/doctors/[id]", "Doctors", "~400", "Detail", "Doctor profile with schedule, appointments", "doctor", "/api/doctors/[id], /api/doctors/[id]/slots", ""],
    # Therapists
    [31, "/therapists", "Therapists", "~400", "List", "Therapist directory", "therapists", "/api/staff?role=therapist", ""],
    [32, "/therapists/new", "Therapists", "~300", "Form", "Add therapist", "form", "/api/staff (POST)", ""],
    [33, "/therapists/[id]", "Therapists", "~400", "Detail", "Therapist profile", "therapist", "/api/staff/[id]", ""],
    # Communications
    [34, "/communications", "Communications", "543", "List/Compose", "Message list with channel filter, compose modal, patient search, template selection", "communications, showCompose, channel, template, patient", "/api/communications, /api/patients, /api/templates", "CommunicationTabs"],
    [35, "/communications/templates", "Communications", "520", "List/CRUD", "Message templates with variable insertion {{var}}, preview, channel/category filters", "templates, showForm, form, categoryFilter", "/api/templates (GET/POST/PUT/DELETE)", "CommunicationTabs"],
    [36, "/communications/bulk", "Communications", "~400", "Form", "Bulk messaging to filtered patient groups", "patients, selectedIds, message, channel", "/api/communications/bulk (POST)", "CommunicationTabs"],
    [37, "/communications/reminders", "Communications", "~400", "List", "Appointment/payment reminders with auto-schedule", "reminders, filter", "/api/reminders, /api/reminders/auto-schedule", "CommunicationTabs"],
    # Treatments
    [38, "/treatments", "Treatments", "~400", "List", "Treatment catalog with categories", "treatments, categoryFilter", "/api/treatments", "TreatmentTabs"],
    [39, "/treatments/plans", "Treatments", "~400", "List", "Treatment plan list", "plans, statusFilter", "/api/treatment-plans", "TreatmentTabs"],
    [40, "/treatments/plans/new", "Treatments", "~500", "Form", "Create treatment plan with items, milestones", "form, items, milestones", "/api/treatment-plans (POST), /api/treatments", "TreatmentTabs"],
    [41, "/treatments/plans/[id]", "Treatments", "~500", "Detail", "Plan detail with progress tracking, milestone updates", "plan, items, milestones, progress", "/api/treatment-plans/[id], /api/treatment-plans/[id]/progress", "TreatmentTabs"],
    [42, "/treatments/progress", "Treatments", "~300", "Dashboard", "Treatment progress overview", "stats", "/api/treatment-plans/stats", "TreatmentTabs"],
    # Packages
    [43, "/packages", "Packages", "~400", "List", "Treatment packages with purchase tracking", "packages", "/api/treatments?type=package", ""],
    [44, "/packages/new", "Packages", "~400", "Form", "Create package with sessions, pricing, validity", "form, treatments", "/api/treatments (POST)", ""],
    [45, "/packages/[id]", "Packages", "~500", "Detail", "Package detail: sessions, sharing, refunds", "pkg, sessions, shares, refunds", "/api/patient-packages/[id], /api/patient-packages/[id]/sessions", ""],
    # Reports
    [46, "/reports", "Reports", "~500", "Dashboard", "Multi-tab reports: Revenue, Inventory, Insurance, Transfers with branch/date filters", "activeTab, reportData, dateRange, branchFilter", "/api/reports, /api/reports/inventory, /api/reports/insurance, /api/reports/transfers", ""],
    # Admin
    [47, "/admin", "Admin", "10", "Redirect", "Redirects to /admin/staff", "", "", ""],
    [48, "/admin/staff", "Admin", "~800", "List/CRUD", "Staff management: CRUD, roles (admin/doctor/therapist/pharmacist), weekly schedule, set password", "staff, showForm, form, schedule, editingId", "/api/staff (GET/POST/PUT), /api/setup-doctor-passwords", "AdminTabs"],
    [49, "/admin/branches", "Admin", "442", "List/CRUD", "Branch management: operating hours, main branch designation", "branches, showForm, form", "/api/branches (GET/POST/PUT)", "AdminTabs"],
    [50, "/admin/audit-log", "Admin", "306", "List", "Audit log with filters (action, entity, user, date range), expandable JSON details", "entries, filters, page", "/api/audit-logs?page=&action=&search=", "AdminTabs"],
    [51, "/admin/settings", "Admin", "~400", "Form", "Clinic settings: name, address, phone, GST, operating hours", "settings, form", "/api/settings (GET/PUT)", "AdminTabs"],
    [52, "/admin/treatments", "Admin", "~400", "List", "Admin treatment catalog management", "treatments", "/api/treatments", "AdminTabs"],
    [53, "/admin/treatments/plans", "Admin", "~400", "List", "Admin treatment plans view", "plans", "/api/treatment-plans", "AdminTabs"],
    [54, "/admin/treatments/plans/new", "Admin", "~400", "Form", "Admin create treatment plan", "form", "/api/treatment-plans (POST)", "AdminTabs"],
    [55, "/admin/treatments/plans/[id]", "Admin", "~400", "Detail", "Admin plan detail/edit", "plan", "/api/treatment-plans/[id]", "AdminTabs"],
    [56, "/admin/treatments/progress", "Admin", "~300", "Dashboard", "Admin treatment progress stats", "stats", "/api/treatment-plans/stats", "AdminTabs"],
    [57, "/admin/users", "Admin", "~400", "List", "User account management", "users", "/api/users", "AdminTabs"],
    # Auth
    [58, "/login", "Auth", "408", "Form", "Multi-step: login → TOTP 2FA → role redirect. Forgot password: email → OTP → reset", "view (login/forgot/reset/totp), email, password, totpCode, otp", "/api/auth/login, /api/auth/totp-verify, /api/auth/forgot-password, /api/auth/reset-password", ""],
    [59, "/invite/[token]", "Auth", "~300", "Form", "Accept invite, set password", "token, password, loading", "/api/invite/[token] (GET/POST)", ""],
    [60, "/security", "Security", "~400", "Page", "2FA setup/disable, password change", "totpStatus, showSetup, qrCode", "/api/auth/totp-setup, /api/auth/totp-verify, /api/auth/totp-disable, /api/auth/totp-status", ""],
]

for i, p in enumerate(pages):
    r = i + 2
    for j, v in enumerate(p): ws1.cell(row=r, column=j+1, value=v)
    row_style(ws1, r, len(cols1), GRAY if i%2==0 else WHITE)

widths1 = [5, 35, 16, 8, 10, 60, 55, 50, 25]
for i, w in enumerate(widths1, 1): ws1.column_dimensions[get_column_letter(i)].width = w
ws1.auto_filter.ref = f"A1:{get_column_letter(len(cols1))}{len(pages)+1}"
ws1.freeze_panes = "A2"

# ════════════════════════════════════════
# SHEET 2: ALL API ROUTES
# ════════════════════════════════════════
ws2 = wb.create_sheet("APIs (108)")
ws2.sheet_properties.tabColor = "1565C0"

cols2 = ["#", "Method", "Endpoint", "Module", "Lines", "Purpose", "Key Logic", "Auth Required"]
for i, h in enumerate(cols2, 1): ws2.cell(row=1, column=i, value=h)
hdr(ws2, 1, len(cols2))

apis = [
    # Patients
    [1, "GET", "/api/patients", "Patients", "84", "List patients with search/filter", "OR search across 6 fields, status filter, includes appointment/communication counts", "Yes"],
    [2, "POST", "/api/patients", "Patients", "144", "Create patient", "NRIC checksum validation, phone normalization (+65/+91), cross-field validation, auto Patient ID, email+WhatsApp notifications", "Yes"],
    [3, "GET", "/api/patients/[id]", "Patients", "29", "Get patient detail", "Includes appointments, communications, clinical notes, documents", "Yes"],
    [4, "PUT", "/api/patients/[id]", "Patients", "80", "Update patient", "NRIC validation, phone normalization, cross-field checks, email lowercase", "Yes"],
    [5, "DELETE", "/api/patients/[id]", "Patients", "20", "Delete patient", "Cascade: communications → appointments → notes → documents → patient", "Yes"],
    [6, "GET", "/api/patients/check-duplicate", "Patients", "130", "Check duplicate NRIC/phone/email", "NRIC: case-insensitive exact. Phone: normalized suffix matching. Email: lowercase exact. Returns matched patient + last visit", "Yes"],
    [7, "GET/POST/PUT/DELETE", "/api/patients/[id]/family", "Patients", "148", "Family member CRUD", "Linked patient support, gender-aware labels, ownership verification", "Yes"],
    [8, "GET", "/api/patients/[id]/timeline", "Patients", "175", "Patient timeline", "Aggregates 6 sources (appointments, notes, docs, comms, vitals, invoices), sorted desc", "Yes"],
    [9, "POST/DELETE", "/api/patients/[id]/photo", "Patients", "106", "Photo upload/delete", "5MB max, image/* only, disk storage in public/uploads/photos/", "Yes"],
    # Appointments
    [10, "GET/POST", "/api/appointments", "Appointments", "~200", "List/create appointments", "Date range filter, doctor filter, status filter, walk-in support", "Yes"],
    [11, "GET/PUT/DELETE", "/api/appointments/[id]", "Appointments", "~150", "Appointment CRUD", "Status workflow: scheduled → confirmed → completed/cancelled/no_show", "Yes"],
    [12, "GET", "/api/appointments/today", "Appointments", "~50", "Today's appointments", "Filtered by today's date range", "Yes"],
    # Inventory
    [13, "GET", "/api/inventory", "Inventory", "84", "List inventory items", "Filters: search, category, subcategory, lowStock, expiringSoon. Includes variants", "Yes"],
    [14, "POST", "/api/inventory", "Inventory", "76", "Create inventory item", "Auto SKU: AYU-[M/H/O/C/E][seq], category validation, 14 fields", "Yes"],
    [15, "GET", "/api/inventory/[id]", "Inventory", "93", "Item detail", "Last 20 transactions, variants, enriched with invoice/PO references", "Yes"],
    [16, "PUT", "/api/inventory/[id]", "Inventory", "88", "Update item", "Auto adjustment transaction if stock changed, records delta", "Yes"],
    [17, "DELETE", "/api/inventory/[id]", "Inventory", "46", "Delete item", "Blocked if transactions exist ('Consider marking as discontinued')", "Yes"],
    [18, "GET/POST", "/api/inventory/[id]/transactions", "Inventory", "289", "Stock transactions", "6 types: purchase/sale/adjustment/return/expired/damaged. Variant-aware. Atomic stock update", "Yes"],
    [19, "GET", "/api/inventory/[id]/movement", "Inventory", "183", "Stock movement graph data", "7/30/60/90 day range, daily in/out/net/closing, peak/lowest, avgDailyOut, daysOfStock", "Yes"],
    [20, "GET", "/api/inventory/alerts", "Inventory", "150", "Low stock + expiry alerts", "6 categories: lowStock, outOfStock, expired, expiringSoon, expiring90, reorderSuggestions. Branch-aware", "Yes"],
    [21, "POST", "/api/inventory/stock-audit", "Inventory", "183", "Submit stock audit", "Bulk: [{itemId, physicalCount}]. Branch-level or global. Returns adjusted/matched/errors", "Yes"],
    [22, "POST", "/api/inventory/import", "Inventory", "127", "CSV bulk import", "Validates category, auto-generates SKU per category. Returns created/failed counts", "Yes"],
    [23, "POST", "/api/inventory/bulk", "Inventory", "~100", "Bulk stock operations", "Bulk adjustment for multiple items", "Yes"],
    [24, "GET", "/api/inventory/stats", "Inventory", "225", "Inventory statistics", "totalItems, totalValue, lowStock, expiring, outOfStock, categoryBreakdown, recentTransactions. Branch-aware", "Yes"],
    [25, "GET", "/api/inventory/lookup", "Inventory", "~50", "Quick item search", "Search by name/SKU/barcode for billing/PO forms", "Yes"],
    [26, "POST", "/api/inventory/expiry-check", "Inventory", "106", "Run expiry check", "Creates notifications for expiring/expired items (deduped within 7 days)", "Yes"],
    [27, "POST", "/api/inventory/write-off-expired", "Inventory", "162", "Write off expired stock", "Branch-specific or global. Creates expired transactions, zeros stock", "Yes"],
    [28, "GET", "/api/inventory/branch-comparison", "Inventory", "132", "Compare stock across branches", "Items × branches matrix, imbalance detection (max > 3× min)", "Yes"],
    [29, "POST", "/api/inventory/indent", "Inventory", "~100", "Internal requisition", "Inter-department stock request", "Yes"],
    # Transfers
    [30, "GET", "/api/transfers", "Transfers", "103", "List transfers", "Filters: status, branchId, search, dateRange. Paginated (20/page)", "Yes"],
    [31, "POST", "/api/transfers", "Transfers", "275", "Create transfer", "Validates branches ≠, stock availability. Auto TRF-YYYYMM-XXXX. Draft status", "Yes"],
    [32, "GET/PUT", "/api/transfers/[id]", "Transfers", "315", "Transfer detail/edit", "Draft-only edit. Replaces items on update. Enriches with user names", "Yes"],
    [33, "POST", "/api/transfers/[id]/submit", "Transfers", "161", "Submit transfer", "draft → in_transit. Atomic: deduct source BranchStock + InventoryItem, create transfer_out txn. Notification to dest", "Yes"],
    [34, "POST", "/api/transfers/[id]/receive", "Transfers", "214", "Receive transfer", "in_transit → received. Accepts partial qty. Atomic: add dest stock, create transfer_in txn. Damaged txn for shortfall", "Yes"],
    [35, "POST", "/api/transfers/[id]/cancel", "Transfers", "170", "Cancel transfer", "Draft: no-op. In_transit: atomic reverse (restore source stock, create reversal txn). Notification", "Yes"],
    [36, "GET/POST", "/api/transfers/templates", "Transfers", "128", "Transfer templates", "CRUD. Enriches with branch names + item details", "Yes"],
    [37, "GET/PUT/DELETE", "/api/transfers/templates/[id]", "Transfers", "~100", "Template detail", "Edit/delete single template", "Yes"],
    # Purchase Orders
    [38, "GET", "/api/purchase-orders", "PO", "73", "List POs", "Filters: status, supplierId, branchId, search. Paginated", "Yes"],
    [39, "POST", "/api/purchase-orders", "PO", "188", "Create PO", "Auto PO-YYYYMM-XXXX. Per-item GST calc. Draft status", "Yes"],
    [40, "GET/PUT/DELETE", "/api/purchase-orders/[id]", "PO", "236", "PO detail/edit/delete", "Draft-only edit/delete. Recalculates totals on item change", "Yes"],
    [41, "POST", "/api/purchase-orders/[id]/receive", "PO", "217", "Receive PO items", "Partial receiving. Atomic: update received qty, increment stock, create purchase txn. Branch-aware (BranchStock upsert)", "Yes"],
    [42, "PUT", "/api/purchase-orders/[id]/status", "PO", "70", "Change PO status", "Allowed: draft→submitted/cancelled, submitted→cancelled/partial, partial→received", "Yes"],
    [43, "GET", "/api/purchase-orders/suggestions", "PO", "220", "Smart reorder suggestions", "90-day usage analysis, avgMonthlyUsage, daysRemaining, suggestedQty, estimatedCost, lastSupplier, fastMovingItems", "Yes"],
    # Suppliers
    [44, "GET/POST", "/api/suppliers", "Suppliers", "~100", "List/create suppliers", "Name, contact, phone, email, GST number, status", "Yes"],
    [45, "GET/PUT", "/api/suppliers/[id]", "Suppliers", "~80", "Supplier detail/edit", "Update contact details, toggle active/inactive", "Yes"],
    # Branches
    [46, "GET/POST", "/api/branches", "Branches", "~120", "List/create branches", "Operating hours, main branch flag", "Yes"],
    [47, "GET/PUT", "/api/branches/[id]", "Branches", "~100", "Branch detail/edit", "Update hours, name, address, phone", "Yes"],
    [48, "GET", "/api/branches/stock", "Branches", "~80", "Branch stock levels", "Stock per item per branch", "Yes"],
    # Billing
    [49, "GET", "/api/invoices", "Billing", "200", "List invoices", "Filters: patientId, status, dateRange, branchId, search. Package sale detection", "Yes"],
    [50, "POST", "/api/invoices", "Billing", "216", "Create invoice", "Auto INV-YYYYMM-XXXX, GST calc, inventory deduction, BranchStock update, StockTransaction audit", "Yes"],
    [51, "GET/PUT/DELETE", "/api/invoices/[id]", "Billing", "226", "Invoice detail/edit/delete", "Paid invoice protection. Stock restore on delete. GST recalc on discount change", "Yes"],
    [52, "POST", "/api/invoices/generate", "Billing", "~100", "Auto-generate invoice", "From appointment or package purchase", "Yes"],
    [53, "GET/POST", "/api/invoices/[id]/payments", "Billing", "146", "Payment CRUD", "Auto REC-YYYYMM-XXXX, validates ≤ balance, atomic status update", "Yes"],
    [54, "GET", "/api/billing/stats", "Billing", "~100", "Billing statistics", "Revenue, outstanding, payment method breakdown", "Yes"],
    # Credit Notes
    [55, "GET/POST", "/api/credit-notes", "Billing", "~120", "Credit note CRUD", "Linked to invoice, pro-rated items", "Yes"],
    [56, "GET", "/api/credit-notes/[id]", "Billing", "~50", "Credit note detail", "Items, linked invoice", "Yes"],
    # Insurance
    [57, "GET/POST", "/api/insurance/providers", "Insurance", "~100", "Provider CRUD", "Corporate/government/private panel types", "Yes"],
    [58, "GET/PUT", "/api/insurance/providers/[id]", "Insurance", "~80", "Provider detail/edit", "", "Yes"],
    [59, "GET/POST", "/api/insurance/claims", "Insurance", "~150", "Claims CRUD", "Pre-auth workflow, status tracking", "Yes"],
    [60, "GET/PUT", "/api/insurance/claims/[id]", "Insurance", "~100", "Claim detail/edit", "Status: pending → approved → settled / rejected", "Yes"],
    # Auth
    [61, "POST", "/api/auth/login", "Auth", "~80", "Login", "Email/password validation, JWT creation (24h), TOTP check", "No"],
    [62, "POST", "/api/auth/logout", "Auth", "~20", "Logout", "Clears JWT cookie", "No"],
    [63, "GET", "/api/auth/me", "Auth", "~30", "Current user", "Verifies JWT, returns user profile", "Yes"],
    [64, "POST", "/api/auth/totp-setup", "Auth", "~60", "Setup 2FA", "Generates TOTP secret + QR code URL", "Yes"],
    [65, "POST", "/api/auth/totp-verify", "Auth", "~50", "Verify TOTP", "Validates 6-digit code, enables 2FA on first verify", "Yes"],
    [66, "GET", "/api/auth/totp-status", "Auth", "~20", "2FA status", "Returns enabled/disabled", "Yes"],
    [67, "POST", "/api/auth/totp-disable", "Auth", "~30", "Disable 2FA", "Clears TOTP secret", "Yes"],
    [68, "POST", "/api/auth/forgot-password", "Auth", "~60", "Forgot password", "Generates OTP, sends via email", "No"],
    [69, "POST", "/api/auth/reset-password", "Auth", "~50", "Reset password", "Validates OTP, sets new password hash", "No"],
    # Communications
    [70, "GET/POST", "/api/communications", "Comms", "~150", "Message CRUD", "Create with channel (email/whatsapp/sms), template variable substitution", "Yes"],
    [71, "POST", "/api/communications/bulk", "Comms", "~100", "Bulk send", "Multiple patients, channel selection, progress tracking", "Yes"],
    [72, "GET/POST/PUT/DELETE", "/api/templates", "Comms", "~200", "Template CRUD", "Variable insertion {{var}}, preview, channel/category", "Yes"],
    [73, "GET/POST", "/api/templates/[id]", "Comms", "~80", "Template detail", "", "Yes"],
    [74, "POST", "/api/templates/preview", "Comms", "~50", "Template preview", "Substitute variables with sample data", "Yes"],
    [75, "GET/POST", "/api/reminders", "Comms", "~120", "Reminder CRUD", "appointment/payment/medication types", "Yes"],
    [76, "PUT/DELETE", "/api/reminders/[id]", "Comms", "~80", "Reminder update/delete", "", "Yes"],
    [77, "POST", "/api/reminders/auto-schedule", "Comms", "~100", "Auto-schedule reminders", "Creates reminders for upcoming appointments", "Yes"],
    [78, "POST", "/api/reminders/bulk", "Comms", "~80", "Bulk reminders", "Send all pending reminders", "Yes"],
    [79, "POST", "/api/reminders/send", "Comms", "~60", "Send single reminder", "Dispatch via configured channel", "Yes"],
    # Clinical
    [80, "GET/POST", "/api/clinical-notes", "Clinical", "~120", "Clinical notes CRUD", "7 types: illness, history, examination, diagnosis, treatment, investigation, follow_up", "Yes"],
    [81, "GET/PUT/DELETE", "/api/clinical-notes/[id]", "Clinical", "~100", "Note detail/edit/delete", "", "Yes"],
    [82, "GET/POST", "/api/prescriptions", "Clinical", "~120", "Prescription CRUD", "Items with dosage, frequency, duration, timing", "Yes"],
    [83, "GET/PUT", "/api/prescriptions/[id]", "Clinical", "~80", "Prescription detail/edit", "", "Yes"],
    [84, "GET/POST", "/api/vitals", "Clinical", "~100", "Vitals CRUD", "BP, pulse, temp, SpO2, weight, height, BMI, blood sugar, respiratory rate", "Yes"],
    [85, "GET/PUT/DELETE", "/api/vitals/[id]", "Clinical", "~80", "Vital detail/edit/delete", "", "Yes"],
    [86, "GET/POST", "/api/documents", "Clinical", "~120", "Document CRUD", "File upload, categories: report/lab/imaging/prescription/consent/other", "Yes"],
    [87, "GET/DELETE", "/api/documents/[id]", "Clinical", "~60", "Document detail/delete", "File cleanup on delete", "Yes"],
    # Treatment Plans
    [88, "GET/POST", "/api/treatment-plans", "Treatments", "~150", "Plan CRUD", "Multi-item plans with estimated sessions/cost", "Yes"],
    [89, "GET/PUT/DELETE", "/api/treatment-plans/[id]", "Treatments", "~120", "Plan detail/edit/delete", "Status: draft → active → completed / cancelled", "Yes"],
    [90, "GET/POST", "/api/treatment-plans/[id]/items", "Treatments", "~100", "Plan items CRUD", "Add/remove treatments from plan", "Yes"],
    [91, "GET/POST", "/api/treatment-plans/[id]/milestones", "Treatments", "~100", "Milestones CRUD", "Goal tracking with target dates", "Yes"],
    [92, "GET/POST", "/api/treatment-plans/[id]/progress", "Treatments", "~80", "Progress tracking", "Session completion, notes", "Yes"],
    [93, "GET", "/api/treatment-plans/stats", "Treatments", "~60", "Plan statistics", "Active, completed, success rate", "Yes"],
    # Packages
    [94, "GET/POST", "/api/patient-packages", "Packages", "~150", "Package CRUD", "Purchase package for patient, sessions, validity", "Yes"],
    [95, "GET", "/api/patient-packages/active", "Packages", "~50", "Active packages", "Patient's current packages with remaining sessions", "Yes"],
    [96, "GET/PUT", "/api/patient-packages/[id]", "Packages", "~100", "Package detail/edit", "", "Yes"],
    [97, "GET/POST", "/api/patient-packages/[id]/sessions", "Packages", "~120", "Session tracking", "Book session, mark complete, remaining count", "Yes"],
    [98, "POST", "/api/patient-packages/[id]/share", "Packages", "~80", "Share package", "Share with family member/linked patient", "Yes"],
    [99, "POST", "/api/patient-packages/[id]/refund", "Packages", "~100", "Process refund", "CaseTrust: cooling-off, 5-day, admin fee calculation", "Yes"],
    # Treatments
    [100, "GET/POST", "/api/treatments", "Treatments", "~120", "Treatment CRUD", "Categories, pricing, duration, package definitions", "Yes"],
    [101, "GET/PUT/DELETE", "/api/treatments/[id]", "Treatments", "~100", "Treatment detail/edit", "", "Yes"],
    [102, "GET", "/api/treatments/[id]/packages", "Treatments", "~50", "Treatment packages", "Package offerings for specific treatment", "Yes"],
    [103, "POST", "/api/treatments/seed", "Treatments", "~80", "Seed treatments", "Pre-populate Ayurvedic treatments", "No"],
    # Reports
    [104, "GET", "/api/reports", "Reports", "~120", "Revenue reports", "Date range, branch filter, summary stats", "Yes"],
    [105, "GET", "/api/reports/inventory", "Reports", "~100", "Inventory reports", "Valuation, movement, expiry, low-stock", "Yes"],
    [106, "GET", "/api/reports/insurance", "Reports", "~80", "Insurance reports", "Claim status, provider breakdown", "Yes"],
    [107, "GET", "/api/reports/transfers", "Reports", "~80", "Transfer reports", "Volume, branch activity, date filters", "Yes"],
    # Admin
    [108, "GET/POST", "/api/staff", "Admin", "~120", "Staff CRUD", "Unified User model, role-based, weekly schedule", "Yes"],
    [109, "GET/PUT/DELETE", "/api/staff/[id]", "Admin", "~100", "Staff detail/edit/delete", "", "Yes"],
    [110, "GET/POST", "/api/users", "Admin", "~100", "User account CRUD", "", "Yes"],
    [111, "GET/PUT", "/api/users/[id]", "Admin", "~80", "User detail/edit", "", "Yes"],
    [112, "GET/PUT", "/api/settings", "Admin", "~80", "Clinic settings", "Singleton config: name, address, GST, hours", "Yes"],
    [113, "GET", "/api/audit-logs", "Admin", "~80", "Audit log", "Paginated, filterable by action/entity/user/date", "Yes"],
    [114, "GET/PUT", "/api/notifications", "Admin", "~80", "Notifications", "Bell icon, mark read, transfer/stock/PO alerts", "Yes"],
    [115, "GET", "/api/dashboard", "Admin", "~100", "Admin dashboard", "Stats: patients, appointments, revenue, upcoming", "Yes"],
    [116, "GET", "/api/doctor/dashboard", "Doctor", "125", "Doctor dashboard", "Today's appointments, upcoming 7d, recent prescriptions, stats", "Yes"],
    [117, "POST", "/api/setup-doctor-passwords", "Admin", "~60", "Set doctor passwords", "Admin sets initial passwords for doctors", "Yes"],
    [118, "GET/POST", "/api/invite/[token]", "Auth", "~100", "Staff invite", "Validate token, set password, activate account", "No"],
]

for i, a in enumerate(apis):
    r = i + 2
    for j, v in enumerate(a): ws2.cell(row=r, column=j+1, value=v)
    row_style(ws2, r, len(cols2), GRAY if i%2==0 else WHITE)
    method_cell = ws2.cell(row=r, column=2)
    m = str(a[1])
    if "POST" in m: method_cell.fill = PatternFill("solid", fgColor=GREEN)
    elif "PUT" in m or "DELETE" in m: method_cell.fill = PatternFill("solid", fgColor=YELLOW)
    method_cell.alignment = center

widths2 = [5, 22, 40, 12, 8, 30, 65, 10]
for i, w in enumerate(widths2, 1): ws2.column_dimensions[get_column_letter(i)].width = w
ws2.auto_filter.ref = f"A1:{get_column_letter(len(cols2))}{len(apis)+1}"
ws2.freeze_panes = "A2"

# ════════════════════════════════════════
# SHEET 3: DATABASE SCHEMA
# ════════════════════════════════════════
ws3 = wb.create_sheet("Schema (41 Models)")
ws3.sheet_properties.tabColor = "E65100"

cols3 = ["#", "Model", "Category", "Fields", "Relations", "Key Fields", "Indexes/Constraints"]
for i, h in enumerate(cols3, 1): ws3.cell(row=1, column=i, value=h)
hdr(ws3, 1, len(cols3))

models = [
    [1, "Patient", "Patients", 30, 11, "patientIdNumber (unique), firstName, lastName, nricId, phone, email, gender, dateOfBirth, age, bloodGroup, ethnicity, nationality, occupation, address, medicalHistory (JSON), groups (JSON), photoUrl, status", "Unique: patientIdNumber"],
    [2, "FamilyMember", "Patients", 7, 1, "patientId, relation, linkedPatientId?, memberName, memberPhone?, memberGender?", "FK: patientId → Patient (cascade)"],
    [3, "Appointment", "Clinical", 18, 2, "patientId, doctorName, date, time, status (scheduled/confirmed/completed/cancelled/no_show), type, reason, department, isWalkIn, notes, packageId?, sessionId?", "FK: patientId → Patient"],
    [4, "Communication", "Clinical", 8, 1, "patientId, type (email/whatsapp/sms), subject?, message, status (sent/failed/pending), sentAt", "FK: patientId → Patient"],
    [5, "ClinicalNote", "Clinical", 9, 1, "patientId, doctorName, type (7 types), title, content, attachments?, createdAt", "FK: patientId → Patient"],
    [6, "Document", "Clinical", 8, 1, "patientId, fileName, fileUrl, category (report/lab/imaging/prescription/consent/other), description?, uploadedBy?, uploadedAt", "FK: patientId → Patient"],
    [7, "Vital", "Clinical", 13, 1, "patientId, date, bloodPressureSystolic, bloodPressureDiastolic, pulse, temperature, oxygenSaturation, weight, height, bmi, bloodSugar, respiratoryRate, notes?", "FK: patientId → Patient"],
    [8, "Treatment", "Treatments", 13, 2, "name, category, description?, duration, price, gstPercent, isActive, type (service/package), sessions?, validityDays?, discount?", ""],
    [9, "TreatmentPackage", "Treatments", 8, 1, "treatmentId, sessions, validityDays, price, discount, isActive", "FK: treatmentId → Treatment"],
    [10, "InventoryItem", "Inventory", 22, 3, "sku (unique), name, category, subcategory?, unit, packing?, manufacturerCode?, costPrice, unitPrice, gstPercent, currentStock, reorderLevel, expiryDate?, batchNumber?, manufacturer?, location?, hsnCode?, description?, status (active/discontinued/out_of_stock)", "Unique: sku"],
    [11, "InventoryVariant", "Inventory", 8, 1, "itemId, variantName, packing, costPrice, unitPrice, gstPercent, sku (unique), currentStock", "FK: itemId → InventoryItem. Unique: sku"],
    [12, "StockTransaction", "Inventory", 12, 1, "itemId, variantId?, type (purchase/sale/adjustment/return/expired/damaged), quantity, unitPrice?, totalAmount?, reference?, notes?, date, batchNumber?", "FK: itemId → InventoryItem"],
    [13, "Supplier", "Inventory", 9, 0, "name, contactPerson?, phone?, email?, address?, gstNumber?, isActive", ""],
    [14, "ClinicSettings", "Admin", 14, 0, "clinicName, address, phone, email, gstNumber, logo?, operatingHours (JSON), currency, taxRate, appointmentDuration, smsEnabled, emailEnabled, whatsappEnabled", "Singleton"],
    [15, "Invoice", "Billing", 20, 3, "invoiceNumber (unique), patientId, patientName, date, dueDate?, subtotal, discountPercent, discountAmount, taxableAmount, gstAmount, totalAmount, paidAmount, balanceAmount, status (draft/pending/partially_paid/paid/cancelled/refunded), notes?, branchId?", "FK: patientId → Patient"],
    [16, "InvoiceItem", "Billing", 12, 1, "invoiceId, description, quantity, unitPrice, discount, gstPercent, gstAmount, totalAmount, itemType?, inventoryItemId?, hsnSacCode?", "FK: invoiceId → Invoice"],
    [17, "Payment", "Billing", 8, 1, "invoiceId, amount, method (cash/card/upi/bank_transfer/insurance), receiptNumber (unique), date, notes?", "FK: invoiceId → Invoice"],
    [18, "InsuranceProvider", "Insurance", 10, 1, "name, panelType (corporate/government/private), contactPerson?, phone?, email?, address?, coverageDetails?, isActive", ""],
    [19, "InsuranceClaim", "Insurance", 14, 2, "patientId, providerId, invoiceId?, claimNumber, amount, approvedAmount?, status (pending/pre_auth/approved/rejected/settled), submittedDate, approvedDate?, settledDate?, notes?", "FKs: patientId, providerId, invoiceId"],
    [20, "CreditNote", "Billing", 10, 1, "invoiceId, creditNoteNumber (unique), amount, reason, items (JSON), date, status", "FK: invoiceId → Invoice"],
    [21, "PurchaseOrder", "PO", 16, 1, "poNumber (unique), supplierId?, supplierName, branchId?, status (draft/submitted/partial/received/cancelled), orderDate, expectedDate?, receivedDate?, subtotal, gstAmount, totalAmount, paidAmount, notes?", "Unique: poNumber"],
    [22, "PurchaseOrderItem", "PO", 9, 1, "purchaseOrderId, inventoryItemId?, itemName, quantity, receivedQty, unitPrice, gstPercent, totalAmount", "FK: purchaseOrderId → PO"],
    [23, "MessageTemplate", "Comms", 8, 0, "name, channel (email/whatsapp/sms), category, subject?, body, variables (JSON), isActive", ""],
    [24, "Reminder", "Comms", 10, 1, "patientId, type (appointment/payment/medication), title, message, scheduledDate, sentDate?, status (pending/sent/failed/cancelled), channel", "FK: patientId → Patient"],
    [25, "Prescription", "Clinical", 8, 2, "patientId, doctorName, date, diagnosis?, notes?, status (active/completed/cancelled)", "FK: patientId → Patient"],
    [26, "PrescriptionItem", "Clinical", 9, 1, "prescriptionId, medicineName, dosage, frequency, duration, timing?, quantity?, notes?", "FK: prescriptionId → Prescription"],
    [27, "TreatmentPlan", "Treatments", 12, 3, "patientId, doctorName, name, description?, status (draft/active/completed/cancelled), startDate, endDate?, totalSessions, completedSessions, estimatedCost, notes?", "FK: patientId → Patient"],
    [28, "TreatmentPlanItem", "Treatments", 9, 1, "planId, treatmentName, sessions, completedSessions, unitPrice, notes?, status (pending/in_progress/completed)", "FK: planId → TreatmentPlan"],
    [29, "TreatmentMilestone", "Treatments", 7, 1, "planId, title, description?, targetDate?, completedDate?, status (pending/achieved/missed)", "FK: planId → TreatmentPlan"],
    [30, "User", "Admin", 22, 0, "email (unique), password, name, role (admin/doctor/therapist/pharmacist/staff), phone?, isActive, totpSecret?, totpEnabled, resetOtp?, resetOtpExpiry?, inviteToken?, inviteExpiry?, specialization?, consultationFee?, qualifications?, schedule (JSON)?", "Unique: email"],
    [31, "Branch", "Admin", 10, 2, "name, address?, phone?, email?, isMainBranch, isActive, operatingHours (JSON)", ""],
    [32, "BranchStock", "Multi-Branch", 6, 2, "branchId, itemId, variantId?, quantity", "Unique: branchId+itemId+variantId"],
    [33, "StockTransfer", "Multi-Branch", 12, 3, "transferNumber (unique), fromBranchId, toBranchId, status (draft/in_transit/received/cancelled), notes?, initiatedBy?, receivedBy?, sentDate?, receivedDate?", "FKs: fromBranch, toBranch"],
    [34, "StockTransferItem", "Multi-Branch", 6, 1, "transferId, itemId, variantId?, quantitySent, quantityReceived", "FK: transferId → StockTransfer"],
    [35, "TransferTemplate", "Multi-Branch", 6, 1, "name, fromBranchId?, toBranchId?, description?, isActive", ""],
    [36, "TransferTemplateItem", "Multi-Branch", 5, 1, "templateId, itemId, variantId?, quantity", "FK: templateId → TransferTemplate"],
    [37, "PatientPackage", "Packages", 14, 4, "patientId, treatmentId, packageName, totalSessions, usedSessions, remainingSessions, price, paidAmount, validFrom, validUntil, status (active/completed/expired/cancelled/refunded), invoiceId?", "FKs: patientId, treatmentId"],
    [38, "PackageSession", "Packages", 9, 1, "packageId, date, doctorName, status (scheduled/completed/cancelled/no_show), notes?, appointmentId?", "FK: packageId → PatientPackage"],
    [39, "PackageShare", "Packages", 6, 1, "packageId, sharedWithPatientId, sharedBy, sessionsAllocated, sessionsUsed", "FKs: packageId, sharedWithPatientId"],
    [40, "PackageRefund", "Packages", 10, 1, "packageId, refundAmount, reason, type (cooling_off/5_day/pro_rated/admin), adminFee?, processedBy, processedDate, notes?", "FK: packageId → PatientPackage"],
    [41, "Notification", "Admin", 8, 0, "type (transfer_incoming/transfer_received/transfer_cancelled/low_stock/po_received/expiry_alert), title, message, link?, isRead, branchId?, createdAt", ""],
    [42, "AuditLog", "Admin", 8, 0, "userId?, userName, action (create/update/delete), entityType, entityId?, changes (JSON)?, ipAddress?, createdAt", ""],
]

for i, m in enumerate(models):
    r = i + 2
    for j, v in enumerate(m): ws3.cell(row=r, column=j+1, value=v)
    row_style(ws3, r, len(cols3), GRAY if i%2==0 else WHITE)

widths3 = [5, 22, 14, 6, 6, 80, 35]
for i, w in enumerate(widths3, 1): ws3.column_dimensions[get_column_letter(i)].width = w
ws3.auto_filter.ref = f"A1:{get_column_letter(len(cols3))}{len(models)+1}"
ws3.freeze_panes = "A2"

# ════════════════════════════════════════
# SHEET 4: COMPONENTS
# ════════════════════════════════════════
ws4 = wb.create_sheet("Components (12)")
ws4.sheet_properties.tabColor = "7B1FA2"

cols4 = ["#", "Component", "File", "Lines", "Purpose", "Props/Interface", "Key Logic", "Used By"]
for i, h in enumerate(cols4, 1): ws4.cell(row=1, column=i, value=h)
hdr(ws4, 1, len(cols4))

components = [
    [1, "LayoutShell", "src/components/LayoutShell.tsx", "37", "Root layout wrapper", "children: ReactNode", "Detects page type (login/invite/doctor), conditionally renders AuthProvider + Sidebar", "Root layout"],
    [2, "Sidebar", "src/components/Sidebar.tsx", "~800", "Main navigation sidebar", "None (reads auth context)", "12 nav items with role-based access, search, alerts bell, reminders, notifications with unread count", "LayoutShell"],
    [3, "AuthProvider", "src/components/AuthProvider.tsx", "92", "Authentication context", "children: ReactNode", "Checks /api/auth/me on mount, redirects to /login if unauthenticated, provides user+logout via context", "LayoutShell"],
    [4, "BarcodeScanner", "src/components/BarcodeScanner.tsx", "199", "Barcode/QR code scanner", "onScan(code), placeholder?, showCamera?", "USB scanner mode (keyboard input) + camera mode (html5-qrcode). Vibration feedback on scan", "Stock audit, PO receiving, Transfer receiving"],
    [5, "StatsCard", "src/components/StatsCard.tsx", "~50", "Dashboard statistics card", "label, value, icon?, color?, trend?", "Colored card with large value, optional trend indicator", "Dashboard, billing, inventory stats"],
    [6, "ConfirmDialog", "src/components/ConfirmDialog.tsx", "~60", "Confirmation modal", "title, message, onConfirm, onCancel, variant? (danger/warning)", "Modal overlay with confirm/cancel buttons, danger variant in red", "Delete actions across all modules"],
    [7, "HelpTip", "src/components/HelpTip.tsx", "~40", "Tooltip/help icon", "text: string", "Info icon with hover tooltip showing help text", "Form pages"],
    [8, "InventoryTabs", "src/components/InventoryTabs.tsx", "~80", "Inventory page navigation", "None (reads pathname)", "Tab bar: Items, Suppliers, PO, Transfers, Branch Stock, Alerts, Audit, Import, Reports", "All 15 inventory pages"],
    [9, "AdminTabs", "src/components/AdminTabs.tsx", "~60", "Admin page navigation", "None (reads pathname)", "Tab bar: Staff, Branches, Settings, Treatments, Audit Log, Users", "All 9 admin pages"],
    [10, "BillingTabs", "src/components/BillingTabs.tsx", "~50", "Billing page navigation", "None (reads pathname)", "Tab bar: Invoices, Insurance, Providers", "All 4 billing pages"],
    [11, "CommunicationTabs", "src/components/CommunicationTabs.tsx", "~50", "Communications page navigation", "None (reads pathname)", "Tab bar: Messages, Templates, Bulk, Reminders", "All 4 communications pages"],
    [12, "TreatmentTabs", "src/components/TreatmentTabs.tsx", "~50", "Treatment page navigation", "None (reads pathname)", "Tab bar: Catalog, Plans, Progress", "All 5 treatment pages"],
]

for i, c in enumerate(components):
    r = i + 2
    for j, v in enumerate(c): ws4.cell(row=r, column=j+1, value=v)
    row_style(ws4, r, len(cols4), GRAY if i%2==0 else WHITE)

widths4 = [5, 20, 38, 8, 30, 35, 55, 35]
for i, w in enumerate(widths4, 1): ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

# ════════════════════════════════════════
# SHEET 5: LIB UTILITIES
# ════════════════════════════════════════
ws5 = wb.create_sheet("Utilities (5)")
ws5.sheet_properties.tabColor = "00695C"

cols5 = ["#", "File", "Lines", "Exports", "Purpose", "Implementation Details"]
for i, h in enumerate(cols5, 1): ws5.cell(row=1, column=i, value=h)
hdr(ws5, 1, len(cols5))

utils = [
    [1, "src/lib/auth.ts", "51", "JWTPayload, createToken(), verifyToken(), hashPassword(), comparePassword(), generateOTP()", "Authentication utilities", "JWT: JOSE library, HS256, 24h expiry. Password: bcryptjs, 12 salt rounds. OTP: 6-digit random (100000-999999)"],
    [2, "src/lib/db.ts", "~15", "prisma (PrismaClient)", "Database connection", "Singleton Prisma client with globalThis caching for dev hot-reload"],
    [3, "src/lib/email.ts", "~60", "sendEmail(to, subject, html)", "Email sending", "Nodemailer transport with SMTP config from env vars. HTML template support"],
    [4, "src/lib/sms.ts", "~40", "sendSMS(to, message)", "SMS sending", "HTTP API integration (configurable provider). Placeholder for production SMS gateway"],
    [5, "src/lib/whatsapp.ts", "~50", "sendWhatsApp(to, message)", "WhatsApp messaging", "WhatsApp Business API integration. Message template support with variable substitution"],
]

for i, u in enumerate(utils):
    r = i + 2
    for j, v in enumerate(u): ws5.cell(row=r, column=j+1, value=v)
    row_style(ws5, r, len(cols5), GRAY if i%2==0 else WHITE)

widths5 = [5, 22, 8, 70, 22, 70]
for i, w in enumerate(widths5, 1): ws5.column_dimensions[get_column_letter(i)].width = w

# ════════════════════════════════════════
# SAVE
# ════════════════════════════════════════
output = "/Users/karthik/Cladue CODE1/patient-registration/docs/Ayur_Centre_Source_Code_Documentation.xlsx"
wb.save(output)
print(f"Saved: {output}")
