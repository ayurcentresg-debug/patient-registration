from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime

wb = Workbook()

# ═══════════════════════════════════════════
# COLOR PALETTE
# ═══════════════════════════════════════════
DARK_GREEN = "1B4332"
MED_GREEN = "2D6A4F"
LIGHT_GREEN = "D8F3DC"
WHITE = "FFFFFF"
LIGHT_GRAY = "F5F5F5"
YELLOW_BG = "FFF9C4"
RED_TEXT = "C62828"
ORANGE_BG = "FFF3E0"
BLUE_BG = "E3F2FD"
GREEN_BG = "E8F5E9"
GRAY_BG = "EEEEEE"
BORDER_COLOR = "BDBDBD"

header_font = Font(name="Arial", bold=True, color=WHITE, size=11)
header_fill = PatternFill("solid", fgColor=DARK_GREEN)
sub_header_font = Font(name="Arial", bold=True, color=DARK_GREEN, size=10)
sub_header_fill = PatternFill("solid", fgColor=LIGHT_GREEN)
body_font = Font(name="Arial", size=10)
body_font_bold = Font(name="Arial", size=10, bold=True)
link_font = Font(name="Arial", size=10, color="1565C0", underline="single")
red_font = Font(name="Arial", size=10, color=RED_TEXT, bold=True)
thin_border = Border(
    left=Side(style="thin", color=BORDER_COLOR),
    right=Side(style="thin", color=BORDER_COLOR),
    top=Side(style="thin", color=BORDER_COLOR),
    bottom=Side(style="thin", color=BORDER_COLOR),
)
wrap_align = Alignment(wrap_text=True, vertical="top")
center_align = Alignment(horizontal="center", vertical="center")
center_wrap = Alignment(horizontal="center", vertical="top", wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_wrap
        cell.border = thin_border

def style_row(ws, row, cols, fill=None):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = body_font
        cell.alignment = wrap_align
        cell.border = thin_border
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)

def auto_width(ws, cols, min_w=12, max_w=40):
    for c in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(c)].width = min_w

# ═══════════════════════════════════════════
# SHEET 1: EXECUTIVE DASHBOARD
# ═══════════════════════════════════════════
ws1 = wb.active
ws1.title = "Dashboard"
ws1.sheet_properties.tabColor = DARK_GREEN

ws1.merge_cells("A1:H1")
title_cell = ws1["A1"]
title_cell.value = "AYUR CENTRE PTE. LTD. — Clinic Management System"
title_cell.font = Font(name="Arial", bold=True, color=DARK_GREEN, size=16)
title_cell.alignment = Alignment(horizontal="center", vertical="center")
ws1.row_dimensions[1].height = 40

ws1.merge_cells("A2:H2")
ws1["A2"].value = f"Project Tracker — Last Updated: {datetime.now().strftime('%d %B %Y')}"
ws1["A2"].font = Font(name="Arial", size=11, color="616161")
ws1["A2"].alignment = Alignment(horizontal="center")

# Summary Stats
stats = [
    ("Total Modules", 12),
    ("Total Pages", 60),
    ("Total API Endpoints", 108),
    ("Database Models", 41),
    ("Days Since Kickoff", "=TODAY()-DATE(2026,3,25)"),
    ("Commits", 44),
]
r = 4
ws1.merge_cells(f"A{r}:H{r}")
ws1[f"A{r}"].value = "PROJECT SUMMARY"
ws1[f"A{r}"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1[f"A{r}"].fill = PatternFill("solid", fgColor=MED_GREEN)
ws1[f"A{r}"].alignment = Alignment(horizontal="center")
ws1.row_dimensions[r].height = 28

r = 5
for i, (label, val) in enumerate(stats):
    col = (i % 3) * 3 + 1
    if i == 3:
        r = 6
    cell_l = ws1.cell(row=r, column=col, value=label)
    cell_l.font = body_font_bold
    cell_l.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    cell_l.border = thin_border
    cell_v = ws1.cell(row=r, column=col + 1, value=val)
    cell_v.font = Font(name="Arial", size=14, bold=True, color=DARK_GREEN)
    cell_v.alignment = center_align
    cell_v.border = thin_border

# Module Status Overview
r = 8
ws1.merge_cells(f"A{r}:H{r}")
ws1[f"A{r}"].value = "MODULE STATUS OVERVIEW"
ws1[f"A{r}"].font = Font(name="Arial", bold=True, color=WHITE, size=12)
ws1[f"A{r}"].fill = PatternFill("solid", fgColor=MED_GREEN)
ws1[f"A{r}"].alignment = Alignment(horizontal="center")

r = 9
cols = ["Module", "Pages", "API Endpoints", "Status", "Completion %", "Tester", "Priority", "Notes"]
for i, h in enumerate(cols, 1):
    ws1.cell(row=r, column=i, value=h)
style_header(ws1, r, 8)

modules = [
    ("Patient Management", 3, 8, "Live", "95%", "", "P0", "NRIC validation, duplicate detection done"),
    ("Appointments", 2, 4, "Live", "90%", "", "P0", "Calendar view, slot management"),
    ("Billing & Invoicing", 4, 10, "Live", "90%", "", "P0", "Invoices, payments, credit notes"),
    ("Inventory Management", 15, 32, "Live", "95%", "", "P0", "Full stock lifecycle, transfers, PO, 45+ sub-features"),
    ("Doctor Portal", 2, 3, "Live", "85%", "", "P1", "Dashboard, prescriptions, clinical notes"),
    ("Treatment Plans", 5, 8, "Live", "85%", "", "P1", "Plans, milestones, progress tracking"),
    ("Packages", 3, 8, "Live", "85%", "", "P1", "Package purchase, sessions, sharing, refunds"),
    ("Communications", 4, 10, "Live", "80%", "", "P2", "Email, WhatsApp, SMS, templates, bulk"),
    ("Insurance", 2, 4, "Live", "75%", "", "P2", "Providers, claims"),
    ("Reports", 1, 5, "Live", "80%", "", "P1", "Revenue, inventory, insurance, transfers"),
    ("Admin & Security", 6, 14, "Live", "90%", "", "P0", "Auth, 2FA, roles, audit log, branches"),
    ("Multi-Branch", 3, 6, "Live", "90%", "", "P0", "Branch stock, transfers, comparison"),
]

for i, mod in enumerate(modules):
    r = 10 + i
    for j, val in enumerate(mod):
        ws1.cell(row=r, column=j + 1, value=val)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws1, r, 8, fill)
    status_cell = ws1.cell(row=r, column=4)
    if mod[3] == "Live":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif mod[3] == "Testing":
        status_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
    status_cell.alignment = center_wrap

ws1.column_dimensions["A"].width = 22
ws1.column_dimensions["B"].width = 10
ws1.column_dimensions["C"].width = 16
ws1.column_dimensions["D"].width = 12
ws1.column_dimensions["E"].width = 14
ws1.column_dimensions["F"].width = 14
ws1.column_dimensions["G"].width = 10
ws1.column_dimensions["H"].width = 45

# ═══════════════════════════════════════════
# SHEET 2: FEATURE TRACKER (Detailed)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("Feature Tracker")
ws2.sheet_properties.tabColor = "2D6A4F"

headers = [
    "Feature ID", "Module", "Feature Name", "What Was Requested",
    "What Was Delivered", "Status", "Date Completed", "Commit Hash",
    "Priority", "Tester", "Test Status", "Bugs Found", "Marketing Tag", "Notes"
]
for i, h in enumerate(headers, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(headers))

features = [
    # Patient Management
    ["F001", "Patients", "Patient Registration Form", "Create patient registration form", "Full registration with 30+ fields, label:input layout, SG address format", "Completed", "2026-03-31", "2a30cc8", "P0", "", "Pending", "", "Patient Onboarding", ""],
    ["F002", "Patients", "Form Redesign — Label:Input Layout", "Redesign form to label-on-left, input-on-right vertical layout", "Flex-based layout with w-40 labels, consistent spacing across all sections", "Completed", "2026-03-31", "3a627ea", "P0", "", "Pending", "", "", "User showed reference screenshot"],
    ["F003", "Patients", "Uniform Field Width (350px)", "Make all fields uniform width, not too long", "fieldMaxWidth=350px constant applied to all input fields globally", "Completed", "2026-03-31", "3a627ea", "P0", "", "Pending", "", "", ""],
    ["F004", "Patients", "NRIC Checksum Validation", "Validate Singapore NRIC structure (9 chars, S/T/F/G/M prefix, checksum)", "Full checksum algorithm, auto-uppercase, real-time validation on blur/change, server-side validation", "Completed", "2026-03-31", "7ca6dcd", "P0", "", "Pending", "", "Singapore Compliance", "Supports S/T (citizens), F/G/M (foreigners)"],
    ["F005", "Patients", "Duplicate Detection — NRIC", "Prevent duplicate NRIC registrations", "Real-time check on blur, hard-block with 'Open Existing Profile' button", "Completed", "2026-03-31", "3a627ea", "P0", "", "Pending", "", "Data Integrity", "Hard block — cannot proceed with duplicate NRIC"],
    ["F006", "Patients", "Duplicate Detection — Phone", "Prevent duplicate phone registrations", "Real-time check with suffix matching, soft warning with override option", "Completed", "2026-03-31", "3a627ea", "P0", "", "Pending", "", "Data Integrity", "Soft block — can override with 'Different person, continue'"],
    ["F007", "Patients", "Duplicate Detection — Email", "Prevent duplicate email registrations", "Real-time check, case-insensitive, soft warning", "Completed", "2026-03-31", "3a627ea", "P1", "", "Pending", "", "", ""],
    ["F008", "Patients", "Phone Normalization on Save", "Normalize phone numbers for consistent storage", "Auto-adds +65 for SG (8-digit), +91 for IN (10-digit), strips formatting", "Completed", "2026-03-31", "7ca6dcd", "P0", "", "Pending", "", "Multi-Country", "Applied to all phone fields: primary, secondary, whatsapp, emergency"],
    ["F009", "Patients", "Cross-Field Phone Validation", "Secondary mobile ≠ primary, emergency ≠ own phone", "Frontend + API validation on both POST and PUT routes", "Completed", "2026-03-31", "7ca6dcd", "P1", "", "Pending", "", "", ""],
    ["F010", "Patients", "Email Lowercase Normalization", "Store emails in consistent format", "Auto-lowercase on save in both POST and PUT routes", "Completed", "2026-03-31", "7ca6dcd", "P1", "", "Pending", "", "", ""],
    ["F011", "Patients", "Patient Detail View", "View full patient profile", "Comprehensive profile with tabs for appointments, notes, documents, vitals", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "", ""],
    ["F012", "Patients", "Patient Photo Upload", "Upload patient photo", "Photo upload API with file storage", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],
    ["F013", "Patients", "Family Members Linking", "Link family members to patients", "Relation dropdown, search existing patients, gender-aware labels", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F014", "Patients", "Patient Timeline", "View patient activity history", "Timeline API showing appointments, notes, communications chronologically", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],

    # Inventory — Core
    ["F020", "Inventory", "Inventory List Page", "Main inventory listing with search/filter", "Paginated list with search, category filter, stock status badges, A-Z sorting", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "Stock Management", "15 pages, 12 inventory pages"],
    ["F021", "Inventory", "Inventory Item Detail Page", "Detailed view of single inventory item", "Sub-tabs: Overview, Transactions, Movement Graph, Stock by Branch", "Completed", "2026-03-30", "2c84221", "P0", "", "Pending", "", "", "Enriched transaction history"],
    ["F022", "Inventory", "Add New Inventory Item", "Form to create inventory items", "Full form with name, category, unit, reorder point, supplier, variants", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", ""],
    ["F023", "Inventory", "Edit Inventory Item", "Update existing inventory item details", "Edit form with all fields, variant management", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", ""],
    ["F024", "Inventory", "Inventory Variants", "Support product variants (sizes, concentrations)", "InventoryVariant model with per-variant pricing, variant picker in billing", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", "Billing shows variant dropdown"],
    ["F025", "Inventory", "Inventory Overhaul — Units & Categories", "Correct units of measure, proper categorization", "Ayurvedic-specific units (ml, g, tablets, sachets), proper categories", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", "Initial data was using wrong units"],
    ["F026", "Inventory", "Transaction History", "Track all stock transactions per item", "StockTransaction model: purchase, sale, adjustment, transfer, write-off", "Completed", "2026-03-30", "2c84221", "P1", "", "Pending", "", "Audit Trail", "API: /api/inventory/[id]/transactions"],
    ["F027", "Inventory", "Stock Movement API", "API endpoint for stock movements", "Records stock in/out with reason, batch, reference", "Completed", "2026-03-30", "2c84221", "P1", "", "Pending", "", "", "API: /api/inventory/[id]/movement"],
    ["F028", "Inventory", "Inventory Statistics Dashboard", "Summary stats for inventory", "Total items, total value, low stock count, expiring count", "Completed", "2026-03-30", "7373b4e", "P1", "", "Pending", "", "Analytics", "API: /api/inventory/stats"],
    ["F029", "Inventory", "Inventory Lookup/Search API", "Quick search for inventory items", "Search by name, SKU, barcode for use in billing/PO forms", "Completed", "2026-03-30", "f572d67", "P1", "", "Pending", "", "", "API: /api/inventory/lookup"],
    ["F030", "Inventory", "Bulk Stock Operations", "Bulk update stock quantities", "Bulk stock adjustment API for multiple items at once", "Completed", "2026-03-30", "7373b4e", "P1", "", "Pending", "", "", "API: /api/inventory/bulk"],

    # Inventory — Import
    ["F031", "Inventory", "CSV/Excel Bulk Import", "Import inventory from CSV/Excel files", "Upload page with drag-drop, column mapping, validation, error reporting, batch create", "Completed", "2026-03-30", "7373b4e", "P1", "", "Pending", "", "Bulk Operations", "API: /api/inventory/import"],

    # Inventory — Stock Audit
    ["F032", "Inventory", "Stock Audit Page", "Physical stock count reconciliation", "Full audit page with barcode scanning, manual entry, discrepancy detection", "Completed", "2026-03-30", "4331cc9", "P0", "", "Pending", "", "Audit Trail", "Barcode scanner integration"],
    ["F033", "Inventory", "Stock Audit API", "Backend for stock audit operations", "Submit audit counts, calculate discrepancies, record adjustments", "Completed", "2026-03-30", "7373b4e", "P0", "", "Pending", "", "", "API: /api/inventory/stock-audit"],
    ["F034", "Inventory", "Barcode/QR Scanning Component", "Scan barcodes for stock operations", "Reusable BarcodeScanner component, used in audit, receiving, and lookup", "Completed", "2026-03-30", "4331cc9", "P1", "", "Pending", "", "Smart Scanning", "src/components/BarcodeScanner.tsx"],

    # Inventory — Suppliers
    ["F035", "Inventory", "Supplier Management Page", "Manage supplier directory", "Supplier list page with CRUD, linked to POs and inventory items", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", ""],
    ["F036", "Inventory", "Supplier CRUD API", "Create/edit/delete suppliers", "Full REST API for supplier management with contact details", "Completed", "2026-03-30", "f572d67", "P0", "", "Pending", "", "", "API: /api/suppliers, /api/suppliers/[id]"],

    # Inventory — Purchase Orders
    ["F037", "Inventory", "Purchase Orders List Page", "View all purchase orders", "Filterable list with status (draft/submitted/partial/received/cancelled)", "Completed", "2026-03-30", "8095ccd", "P0", "", "Pending", "", "Procurement", ""],
    ["F038", "Inventory", "Create Purchase Order Page", "Form to create new PO", "Select supplier, add line items, quantities, prices, branch selection", "Completed", "2026-03-30", "8095ccd", "P0", "", "Pending", "", "", "Branch-aware since 7b18ad9"],
    ["F039", "Inventory", "Purchase Order Detail Page", "View/manage single PO", "Full PO view with status actions, line items, receive button", "Completed", "2026-03-30", "8095ccd", "P0", "", "Pending", "", "", ""],
    ["F040a", "Inventory", "PO Status Workflow", "Draft → Submitted → Partial → Received", "Status change API with validation rules", "Completed", "2026-03-30", "8095ccd", "P0", "", "Pending", "", "", "API: /api/purchase-orders/[id]/status"],
    ["F040b", "Inventory", "PO Receiving Workflow", "Receive goods against POs", "Partial receiving, quantity validation, auto-stock-update on receive", "Completed", "2026-03-30", "8095ccd", "P0", "", "Pending", "", "", "API: /api/purchase-orders/[id]/receive"],
    ["F040c", "Inventory", "Smart PO Suggestions", "Auto-suggest items needing reorder", "Algorithm based on current stock vs reorder point, consumption rate, fast-moving items", "Completed", "2026-03-30", "8095ccd", "P1", "", "Pending", "", "AI-Powered", "API: /api/purchase-orders/suggestions"],
    ["F040d", "Inventory", "Branch-Aware Purchase Orders", "POs linked to specific branch", "Branch selector on PO, received stock goes to correct branch", "Completed", "2026-03-30", "7b18ad9", "P0", "", "Pending", "", "Multi-Branch", ""],

    # Inventory — Stock Transfers
    ["F041a", "Inventory", "Stock Transfers List Page", "View all inter-branch transfers", "Filterable list by status, source/destination branch, date", "Completed", "2026-03-30", "dadee2c", "P0", "", "Pending", "", "Multi-Branch", ""],
    ["F041b", "Inventory", "Create Transfer Page", "Form to initiate stock transfer", "Select source/dest branch, add items with quantities, notes", "Completed", "2026-03-30", "dadee2c", "P0", "", "Pending", "", "", ""],
    ["F041c", "Inventory", "Transfer Detail Page", "View/manage single transfer", "Full view with items, status actions (submit/receive/cancel)", "Completed", "2026-03-30", "dadee2c", "P0", "", "Pending", "", "", "Fixed: 615e22f, e2ba711"],
    ["F041d", "Inventory", "Transfer Submit/Receive/Cancel APIs", "Complete transfer lifecycle", "Submit (deduct source stock), Receive (add dest stock), Cancel (reverse)", "Completed", "2026-03-30", "dadee2c", "P0", "", "Pending", "", "", "4 API routes for transfer actions"],
    ["F041e", "Inventory", "Transfer Templates", "Save and reuse frequent transfers", "Template CRUD, apply template to pre-fill new transfer items", "Completed", "2026-03-31", "91d2fb3", "P2", "", "Pending", "", "Efficiency", "API: /api/transfers/templates"],
    ["F041f", "Inventory", "Transfer Notifications", "Notify branches of incoming transfers", "Notification model with bell icon in header, unread count badge", "Completed", "2026-03-30", "7e37183", "P1", "", "Pending", "", "", "API: /api/notifications"],

    # Inventory — Branch Stock
    ["F042a", "Inventory", "Branch Stock Page", "View stock levels per branch", "Branch-specific stock listing with totals and summary cards", "Completed", "2026-03-30", "c853141", "P0", "", "Pending", "", "Multi-Branch", ""],
    ["F042b", "Inventory", "Multi-Branch Stock Comparison", "Side-by-side stock comparison across branches", "Comparison view with filtering, highlights low stock per branch", "Completed", "2026-03-30", "95a6f8d", "P1", "", "Pending", "", "Multi-Branch", "API: /api/inventory/branch-comparison"],
    ["F042c", "Inventory", "Branch Stock API", "Backend for branch-level stock queries", "Stock levels, branch totals, per-item per-branch breakdown", "Completed", "2026-03-30", "c853141", "P0", "", "Pending", "", "", "API: /api/branches/stock"],

    # Inventory — Alerts & Expiry
    ["F043a", "Inventory", "Alerts Dashboard Page", "Central alert management", "Combined view of low stock alerts + expiry alerts with action buttons", "Completed", "2026-03-30", "4ba2d4a", "P0", "", "Pending", "", "Compliance", ""],
    ["F043b", "Inventory", "Low Stock Alerts", "Alert when stock falls below reorder point", "Alert API + badge counts + doctor portal card showing low stock items", "Completed", "2026-03-30", "616ffff", "P0", "", "Pending", "", "", "API: /api/inventory/alerts"],
    ["F043c", "Inventory", "Expiry Management & Auto-Alerts", "Track and alert on expiring items", "Auto-check for items expiring within 30/60/90 days, write-off workflow", "Completed", "2026-03-30", "4ba2d4a", "P0", "", "Pending", "", "Compliance", "API: /api/inventory/expiry-check"],
    ["F043d", "Inventory", "Expired Stock Write-Off", "Write off expired stock per branch", "Branch-level write-off with reason, auto-deduct stock, transaction record", "Completed", "2026-03-30", "4ba2d4a", "P1", "", "Pending", "", "", "API: /api/inventory/write-off-expired"],

    # Inventory — Analytics
    ["F044a", "Inventory", "Stock Movement History Graph", "Visualize stock movement trends", "Interactive chart on item detail page showing stock over time", "Completed", "2026-03-30", "dff7926", "P2", "", "Pending", "", "Analytics", ""],
    ["F044b", "Inventory", "Inventory Reports Page", "Comprehensive stock analytics", "Valuation report, movement report, expiry report, low-stock report with date filters", "Completed", "2026-03-30", "7373b4e", "P1", "", "Pending", "", "Analytics", "API: /api/reports/inventory"],
    ["F044c", "Inventory", "Transfer Reports", "Report on inter-branch transfer activity", "Transfer volume, branch activity, date range and branch filters", "Completed", "2026-03-30", "004b1e2", "P2", "", "Pending", "", "", "API: /api/reports/transfers"],
    ["F044d", "Inventory", "Indent/Requisition API", "Internal stock request between departments", "Indent request creation and approval workflow", "Completed", "2026-03-30", "7373b4e", "P2", "", "Pending", "", "", "API: /api/inventory/indent"],

    # Inventory — UI Polish
    ["F045a", "Inventory", "InventoryTabs Navigation Component", "Consistent tab navigation across all inventory pages", "Reusable InventoryTabs component with active state highlighting", "Completed", "2026-03-31", "a9fafd6", "P1", "", "Pending", "", "", "Fixed duplicate tabs bug"],
    ["F045b", "Inventory", "Header Consistency Fix", "Consistent headers across inventory sub-pages", "Unified page headers, back buttons, padding, grid layout", "Completed", "2026-03-31", "2b69c86", "P1", "", "Pending", "", "", ""],
    ["F045c", "Inventory", "Branch-Stock Card Colors Fix", "Summary cards using correct YODA design tokens", "Fixed hardcoded colors to use CSS variables from design system", "Completed", "2026-03-31", "e08d700", "P1", "", "Pending", "", "", ""],
    ["F045d", "Inventory", "Navigation Consistency Fix", "Tabs, back buttons, padding across all pages", "Unified navigation patterns, consistent spacing", "Completed", "2026-03-31", "e307402", "P1", "", "Pending", "", "", ""],
    ["F045e", "Inventory", "A-Z Sorting Fix", "Inventory sorted alphabetically by default", "Fixed sorting across all inventory list endpoints", "Completed", "2026-03-30", "ef27e26", "P1", "", "Pending", "", "", ""],
    ["F045f", "Inventory", "Inventory Seed Data", "Pre-populated inventory for Railway", "80+ Ayurvedic products seeded with variants, prices, stock levels", "Completed", "2026-03-30", "c130bdc", "P0", "", "Pending", "", "", ""],
    ["F045g", "Inventory", "End-to-End Bug Fixes", "6 bugs found in code audit", "Fixed transfer receive mapping, detail page data, and other issues", "Completed", "2026-03-30", "acfd45e", "P0", "", "Pending", "", "", "Commits: e2ba711, 615e22f, acfd45e"],

    # Inventory — Branch Awareness
    ["F046a", "Inventory", "Branch-Aware Dashboard", "Dashboard shows branch-specific data", "Dashboard stats filtered by selected branch", "Completed", "2026-03-30", "c853141", "P0", "", "Pending", "", "Multi-Branch", ""],
    ["F046b", "Inventory", "Branch-Aware Billing", "Billing filtered by branch", "Branch dropdown filter on billing list page", "Completed", "2026-03-30", "685785a", "P1", "", "Pending", "", "Multi-Branch", ""],
    ["F046c", "Inventory", "Branch-Aware Reports", "Reports filtered by branch", "Branch filter on revenue, inventory, and transfer reports", "Completed", "2026-03-30", "c853141", "P1", "", "Pending", "", "Multi-Branch", ""],
    ["F046d", "Inventory", "Branch-Aware Stock Audit", "Audit stock per branch", "Branch selection in stock audit page", "Completed", "2026-03-30", "c853141", "P0", "", "Pending", "", "Multi-Branch", ""],
    ["F046e", "Inventory", "Low Stock Alert in Doctor Portal", "Doctors see low stock for their branch", "Low stock card on doctor dashboard showing critical items", "Completed", "2026-03-30", "616ffff", "P1", "", "Pending", "", "", ""],

    # Billing
    ["F040", "Billing", "Invoice Generation", "Create invoices for services", "Auto-generate invoices with line items, tax, discounts", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "", ""],
    ["F041", "Billing", "Payment Recording", "Record payments against invoices", "Multiple payment methods, partial payments, payment history", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "", ""],
    ["F042", "Billing", "Credit Notes", "Issue refunds/credits", "Credit note creation, link to invoice", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F043", "Billing", "Branch Filter", "Filter billing by branch", "Branch dropdown filter on billing list page", "Completed", "2026-03-30", "685785a", "P1", "", "Pending", "", "Multi-Branch", ""],
    ["F044", "Billing", "Billing Statistics", "Dashboard stats for billing", "Revenue stats API with period filters", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],

    # Doctor Portal
    ["F050", "Doctor Portal", "Doctor Dashboard", "Dedicated doctor view", "Today's appointments, patients, low stock alerts", "Completed", "2026-03-29", "11d63fa", "P0", "", "Pending", "", "Role-Based Access", ""],
    ["F051", "Doctor Portal", "Prescriptions", "Write and manage prescriptions", "Prescription CRUD with items, dosage, duration", "Completed", "2026-03-29", "11d63fa", "P0", "", "Pending", "", "", ""],
    ["F052", "Doctor Portal", "Clinical Notes", "Add clinical notes to patient visits", "Rich text notes with attachments", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "", ""],

    # Auth & Admin
    ["F060", "Admin", "Authentication System", "Secure login", "JWT-based auth with role-based access (admin/doctor/staff)", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "Security", ""],
    ["F061", "Admin", "Two-Factor Authentication (2FA)", "TOTP-based 2FA", "Setup, verify, disable TOTP with QR code generation", "Completed", "2026-03-28", "3fd92f7", "P0", "", "Pending", "", "Security", ""],
    ["F062", "Admin", "Staff Management", "Manage clinic staff", "Unified User model with roles, invite system, password management", "Completed", "2026-03-28", "b38c37c", "P0", "", "Pending", "", "", ""],
    ["F063", "Admin", "Audit Log", "Track all system changes", "AuditLog model with user, action, timestamp, details", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "Compliance", ""],
    ["F064", "Admin", "Branch Management", "Multi-branch support", "Branch CRUD, branch-level stock, branch-aware operations", "Completed", "2026-03-30", "c853141", "P0", "", "Pending", "", "Multi-Branch", ""],
    ["F065", "Admin", "Branding — Ayur Centre", "Rebrand from PatientReg", "Logo, app name, sidebar, all pages updated", "Completed", "2026-03-28", "1fa296b", "P0", "", "Pending", "", "", ""],

    # Communications
    ["F070", "Communications", "Email Integration", "Send emails to patients", "Email sending with template support", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F071", "Communications", "WhatsApp Integration", "Send WhatsApp messages", "WhatsApp API integration", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F072", "Communications", "Message Templates", "Reusable message templates", "Template CRUD with variable substitution, preview", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F073", "Communications", "Bulk Messaging", "Send messages to multiple patients", "Bulk send with filtering, progress tracking", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],
    ["F074", "Communications", "Appointment Reminders", "Auto-send reminders", "Auto-schedule, manual send, bulk reminders", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "Automation", ""],

    # Packages & Treatments
    ["F080", "Packages", "Package Management", "Create treatment packages", "Packages with sessions, pricing, validity", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F081", "Packages", "Session Tracking", "Track package session usage", "Session booking, completion, remaining count", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "", ""],
    ["F082", "Packages", "Package Sharing", "Share packages between family members", "Share API with linked patients", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],
    ["F083", "Packages", "Package Refunds", "Process package refunds", "Pro-rated refund calculation", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],

    # Treatment Plans
    ["F085", "Treatment Plans", "Treatment Plan Builder", "Create structured treatment plans", "Multi-item plans with milestones and progress tracking", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "Clinical Excellence", ""],

    # Reports
    ["F090", "Reports", "Revenue Reports", "Revenue analytics", "Date range, branch filter, summary stats", "Completed", "2026-03-28", "3fd92f7", "P1", "", "Pending", "", "Analytics", ""],
    ["F091", "Reports", "Inventory Reports", "Stock analytics", "Valuation, movement, expiry, low-stock reports", "Completed", "2026-03-30", "7373b4e", "P1", "", "Pending", "", "", ""],
    ["F092", "Reports", "Insurance Reports", "Insurance claim analytics", "Claim status, provider breakdown", "Completed", "2026-03-28", "3fd92f7", "P2", "", "Pending", "", "", ""],
    ["F093", "Reports", "Transfer Reports", "Stock transfer analytics", "Transfer volume, branch activity", "Completed", "2026-03-30", "004b1e2", "P2", "", "Pending", "", "", ""],

    # Deployment
    ["F100", "DevOps", "Railway Deployment", "Deploy to Railway with persistent storage", "SQLite on persistent volume, auto-seed on deploy", "Completed", "2026-03-29", "001c9f9", "P0", "", "Pending", "", "Cloud Hosted", ""],
    ["F101", "DevOps", "Data Seeding", "Pre-populate with test data", "Doctors, therapists, patients, inventory all seeded", "Completed", "2026-03-30", "2c500ae", "P0", "", "Pending", "", "", ""],

    # Pending
    ["F110", "Patients", "Multi-Country ID Selector", "Dropdown for ID type (NRIC/Aadhaar/Passport)", "Analysis completed, code pending", "Planned", "", "", "P1", "", "", "", "Multi-Country", "Part of localization roadmap"],
    ["F111", "Patients", "Address Structure per Country", "SG: Block/Street/Unit/Postal | IN: different format", "Analysis completed, code pending", "Planned", "", "", "P1", "", "", "", "Multi-Country", ""],
    ["F112", "Patients", "Database Unique Constraints", "Add partial unique indexes for NRIC, phone, email", "Analysis completed, code pending", "Planned", "", "", "P1", "", "", "", "Data Integrity", "Step 4 of duplicate prevention"],
    ["F113", "Patients", "Duplicate Merge Tool", "Admin tool to merge duplicate patient records", "Analysis completed, code pending", "Planned", "", "", "P2", "", "", "", "", "Step 5 of duplicate prevention"],
]

for i, feat in enumerate(features):
    r = i + 2
    for j, val in enumerate(feat):
        ws2.cell(row=r, column=j + 1, value=val)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws2, r, len(headers), fill)

    # Color-code status
    status_cell = ws2.cell(row=r, column=6)
    status = feat[5]
    if status == "Completed":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif status == "Testing":
        status_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
    elif status == "Planned":
        status_cell.fill = PatternFill("solid", fgColor=BLUE_BG)
    elif status == "In Progress":
        status_cell.fill = PatternFill("solid", fgColor=ORANGE_BG)
    status_cell.alignment = center_wrap

    # Color-code test status
    test_cell = ws2.cell(row=r, column=11)
    if feat[10] == "Passed":
        test_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif feat[10] == "Failed":
        test_cell.fill = PatternFill("solid", fgColor="FFCDD2")
    elif feat[10] == "Pending":
        test_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
    test_cell.alignment = center_wrap

col_widths = [10, 14, 30, 40, 50, 12, 14, 12, 10, 14, 12, 10, 16, 35]
for i, w in enumerate(col_widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(features) + 1}"
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════
# SHEET 3: VALIDATION CHECKLIST
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("Validation Checklist")
ws3.sheet_properties.tabColor = "E65100"

val_headers = ["Field", "Module", "Validation Rule", "Frontend?", "API?", "Country", "Status", "Tested By", "Test Date", "Notes"]
for i, h in enumerate(val_headers, 1):
    ws3.cell(row=1, column=i, value=h)
style_header(ws3, 1, len(val_headers))

validations = [
    ["First Name", "Patients", "Required, trimmed, non-empty", "Yes", "Yes", "All", "Done", "", "", "Unicode, hyphens, apostrophes accepted"],
    ["Last Name", "Patients", "Required, trimmed, non-empty", "Yes", "Yes", "All", "Done", "", "", "Should become optional for India (mononyms)"],
    ["NRIC ID", "Patients", "9 chars: S/T/F/G/M + 7 digits + checksum letter", "Yes", "Yes", "SG", "Done", "", "", "Auto-uppercase, checksum algorithm validated"],
    ["NRIC Duplicate", "Patients", "Hard block — cannot save if NRIC already exists", "Yes", "Yes", "SG", "Done", "", "", "Shows existing patient link"],
    ["Phone (Primary)", "Patients", "Required, SG 8-digit / IN 10-digit / International +XX", "Yes", "Yes", "SG/IN", "Done", "", "", "Normalized to +65/+91 on save"],
    ["Phone Duplicate", "Patients", "Soft warning — can override with 'Different person'", "Yes", "Yes", "All", "Done", "", "", "Suffix matching for country code differences"],
    ["Phone Normalization", "Patients", "Strip formatting, add +65 (SG) or +91 (IN) prefix", "No", "Yes", "SG/IN", "Done", "", "", "Applied to all phone fields on save"],
    ["Secondary Mobile", "Patients", "Optional, valid phone format, ≠ primary mobile", "Yes", "Yes", "All", "Done", "", "", "Cross-field validation"],
    ["Emergency Phone", "Patients", "Optional, valid phone, ≠ patient's own number", "Yes", "Yes", "All", "Done", "", "", "Cross-field validation"],
    ["Email", "Patients", "Optional, valid format, case-insensitive duplicate check", "Yes", "Yes", "All", "Done", "", "", "Lowercased on save"],
    ["Email Duplicate", "Patients", "Soft warning — can override", "Yes", "No", "All", "Done", "", "", ""],
    ["Gender", "Patients", "Required, must be male/female/other", "Yes", "Yes", "All", "Done", "", "", "Radio buttons"],
    ["Date of Birth", "Patients", "Optional, valid date or age entry", "Yes", "No", "All", "Done", "", "", "DOB or age toggle"],
    ["Aadhaar (India)", "Patients", "12 digits with Verhoeff checksum", "No", "No", "IN", "Planned", "", "", "Part of multi-country ID support"],
    ["Passport", "Patients", "Alphanumeric, country-specific format", "No", "No", "All", "Planned", "", "", "Part of multi-country ID support"],
]

for i, val in enumerate(validations):
    r = i + 2
    for j, v in enumerate(val):
        ws3.cell(row=r, column=j + 1, value=v)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws3, r, len(val_headers), fill)
    status_cell = ws3.cell(row=r, column=7)
    if val[6] == "Done":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif val[6] == "Planned":
        status_cell.fill = PatternFill("solid", fgColor=BLUE_BG)
    status_cell.alignment = center_wrap

val_widths = [18, 12, 50, 10, 10, 10, 10, 12, 12, 45]
for i, w in enumerate(val_widths, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.auto_filter.ref = f"A1:{get_column_letter(len(val_headers))}{len(validations) + 1}"
ws3.freeze_panes = "A2"

# ═══════════════════════════════════════════
# SHEET 4: GIT HISTORY
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("Git History")
ws4.sheet_properties.tabColor = "424242"

git_headers = ["Commit", "Date", "Description", "Module", "Type"]
for i, h in enumerate(git_headers, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(git_headers))

commits = [
    ["7ca6dcd", "2026-03-31", "Add NRIC checksum validation, phone normalization, and cross-field checks", "Patients", "Feature"],
    ["3a627ea", "2026-03-31", "Redesign patient form: label-input layout, 350px field width, duplicate detection", "Patients", "Feature"],
    ["0ffb9b3", "2026-03-31", "Reorder patient registration form to match patient detail view", "Patients", "Enhancement"],
    ["4387814", "2026-03-31", "Fix patient form layout: compact fields, consistent spacing", "Patients", "Fix"],
    ["2a30cc8", "2026-03-31", "Redesign patient registration form: layout, dropdowns, SG address", "Patients", "Feature"],
    ["e08d700", "2026-03-31", "Fix branch-stock summary card colors to use YODA tokens", "Inventory", "Fix"],
    ["a9fafd6", "2026-03-31", "Fix duplicate InventoryTabs on branch-stock page", "Inventory", "Fix"],
    ["2b69c86", "2026-03-31", "Fix header consistency across all inventory sub-pages", "Inventory", "Fix"],
    ["e307402", "2026-03-31", "Fix navigation consistency: tabs, back buttons, padding, grid", "UI/UX", "Fix"],
    ["91d2fb3", "2026-03-31", "Add transfer templates: save and reuse frequent transfers", "Inventory", "Feature"],
    ["95a6f8d", "2026-03-30", "Add multi-branch stock comparison view", "Inventory", "Feature"],
    ["dff7926", "2026-03-30", "Add stock movement history graph to inventory item detail", "Inventory", "Feature"],
    ["685785a", "2026-03-30", "Add branch filter to billing list", "Billing", "Feature"],
    ["4ba2d4a", "2026-03-30", "Add expiry management: auto-alerts, branch write-off, enhanced alerts", "Inventory", "Feature"],
    ["acfd45e", "2026-03-30", "Fix 6 bugs from end-to-end code audit", "Multiple", "Fix"],
    ["e2ba711", "2026-03-30", "Fix transfer receive: map transfer item ID to inventory item ID", "Inventory", "Fix"],
    ["615e22f", "2026-03-30", "Fix transfer detail page: map nested item data to flat structure", "Inventory", "Fix"],
    ["4331cc9", "2026-03-30", "Add barcode/QR scanning for stock audit and receiving", "Inventory", "Feature"],
    ["7e37183", "2026-03-30", "Add notification system for transfer alerts", "Inventory", "Feature"],
    ["004b1e2", "2026-03-30", "Add transfer reports to inventory and main reports pages", "Reports", "Feature"],
    ["616ffff", "2026-03-30", "Add low stock alert card to doctor portal dashboard", "Doctor Portal", "Feature"],
    ["7b18ad9", "2026-03-30", "Add branch awareness to Purchase Orders", "Inventory", "Feature"],
    ["ef27e26", "2026-03-30", "Fix inventory sorting to A-Z by name across all endpoints", "Inventory", "Fix"],
    ["c853141", "2026-03-30", "Add branch awareness across billing, reports, dashboard, and stock audit", "Multi-Branch", "Feature"],
    ["dadee2c", "2026-03-30", "Add stock transfers between branches feature", "Inventory", "Feature"],
    ["2c84221", "2026-03-30", "Add sub-tabs and enriched transaction history on inventory detail page", "Inventory", "Enhancement"],
    ["8095ccd", "2026-03-30", "Add smart PO suggestions, receiving workflow, and fast-moving items", "Inventory", "Feature"],
    ["7373b4e", "2026-03-30", "Add P1+P2 inventory features: import, stock audit, reports, indent", "Inventory", "Feature"],
    ["2c500ae", "2026-03-30", "Seed all data to Railway: patients, appointments, invoices, treatments", "DevOps", "Enhancement"],
    ["c130bdc", "2026-03-30", "Add inventory seed for Railway + fix build error", "DevOps", "Fix"],
    ["f572d67", "2026-03-30", "Inventory overhaul: correct units, variant support, billing variant picker", "Inventory", "Feature"],
    ["1baa2e2", "2026-03-29", "Move db push and seed to start script for Railway volume", "DevOps", "Fix"],
    ["d9d858c", "2026-03-29", "Trigger redeploy for Railway volume support", "DevOps", "Fix"],
    ["1d94f95", "2026-03-29", "Fix build errors preventing Railway deployment", "DevOps", "Fix"],
    ["16eb698", "2026-03-29", "Fix DB path: simplify to always use dev.db unless DB_PATH is set", "DevOps", "Fix"],
    ["bbf8671", "2026-03-29", "Seed all doctors, therapists, and patients on every deploy", "DevOps", "Enhancement"],
    ["001c9f9", "2026-03-29", "Fix data loss on Railway: use persistent volume for SQLite database", "DevOps", "Fix"],
    ["5933f75", "2026-03-29", "Add Set Password button in Admin Staff page for managing doctor logins", "Admin", "Feature"],
    ["11d63fa", "2026-03-29", "Add doctor portal, prescriptions, PDF export, role-based access, and staff password management", "Doctor Portal", "Feature"],
    ["b38c37c", "2026-03-28", "Unify staff management: merge Doctor model into User with role-based access", "Admin", "Feature"],
    ["1fa296b", "2026-03-28", "Rebrand app from PatientReg to Ayur Centre Pte. Ltd.", "Admin", "Enhancement"],
    ["7dee12a", "2026-03-28", "Add admin seed to build script for Railway deployment", "DevOps", "Enhancement"],
    ["3fd92f7", "2026-03-28", "Clinic management system with auth, 2FA, and full features", "All", "Feature"],
    ["6bc9deb", "2026-03-25", "Initial commit from Create Next App", "DevOps", "Init"],
]

for i, c in enumerate(commits):
    r = i + 2
    for j, v in enumerate(c):
        ws4.cell(row=r, column=j + 1, value=v)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws4, r, len(git_headers), fill)
    type_cell = ws4.cell(row=r, column=5)
    t = c[4]
    if t == "Feature":
        type_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif t == "Fix":
        type_cell.fill = PatternFill("solid", fgColor="FFCDD2")
    elif t == "Enhancement":
        type_cell.fill = PatternFill("solid", fgColor=BLUE_BG)
    type_cell.alignment = center_wrap

git_widths = [12, 14, 70, 16, 14]
for i, w in enumerate(git_widths, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.auto_filter.ref = f"A1:{get_column_letter(len(git_headers))}{len(commits) + 1}"
ws4.freeze_panes = "A2"

# ═══════════════════════════════════════════
# SHEET 5: MARKETING FEATURES
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("Marketing Copy")
ws5.sheet_properties.tabColor = "1565C0"

mkt_headers = ["Category", "Feature Headline", "One-Liner Description", "Key Benefit", "Target Audience", "Screenshot Needed?"]
for i, h in enumerate(mkt_headers, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(mkt_headers))

marketing = [
    ["Patient Management", "Smart Patient Registration", "30+ field registration with NRIC validation, duplicate detection, and auto-normalization", "Zero duplicate records, Singapore-compliant", "Clinic Admin", "Yes"],
    ["Patient Management", "Real-Time Duplicate Prevention", "Instant alerts when NRIC, phone, or email matches an existing patient", "Prevents data entry errors before they happen", "Clinic Admin", "Yes"],
    ["Patient Management", "Singapore NRIC Checksum Validation", "Built-in NRIC/FIN checksum verification catches typos instantly", "Singapore regulatory compliance", "Clinic Owner", "Yes"],
    ["Inventory", "End-to-End Inventory Management", "From purchase orders to stock audit, with barcode scanning and expiry tracking", "Complete visibility into stock at every branch", "Operations Manager", "Yes"],
    ["Inventory", "Multi-Branch Stock Transfers", "Transfer stock between branches with templates, notifications, and full tracking", "Optimized stock distribution across locations", "Operations Manager", "Yes"],
    ["Inventory", "Smart Reorder Suggestions", "AI-powered purchase order suggestions based on consumption patterns", "Never run out of critical supplies", "Procurement", "Yes"],
    ["Inventory", "Barcode & QR Scanning", "Scan products during stock audit and receiving for error-free operations", "Faster, more accurate stock counts", "Warehouse Staff", "Yes"],
    ["Inventory", "Expiry Alert Dashboard", "Automated tracking and branch-level write-off for expiring products", "Reduce waste, ensure patient safety", "Quality Manager", "Yes"],
    ["Doctor Portal", "Dedicated Doctor Dashboard", "Personalized view with today's patients, prescriptions, and alerts", "Doctors see only what matters to them", "Doctors", "Yes"],
    ["Billing", "Comprehensive Billing Suite", "Invoicing, payments, credit notes, and insurance claims in one place", "Streamlined revenue collection", "Billing Staff", "Yes"],
    ["Security", "Enterprise-Grade Security", "JWT authentication, TOTP 2FA, role-based access, and complete audit trail", "PDPA-compliant data protection", "Clinic Owner", "Yes"],
    ["Multi-Branch", "Multi-Branch Operations", "Manage inventory, billing, and staff across all your clinic locations", "Scale your practice without scaling complexity", "Clinic Owner", "Yes"],
    ["Communications", "Patient Communication Hub", "Email, WhatsApp, SMS with templates, bulk messaging, and auto-reminders", "Reduce no-shows and improve engagement", "Clinic Admin", "Yes"],
    ["Packages", "Treatment Package Management", "Sell packages, track sessions, share between family members, process refunds", "Increase revenue through package offerings", "Clinic Owner", "Yes"],
    ["Analytics", "Business Intelligence Reports", "Revenue, inventory, insurance, and transfer analytics with branch breakdowns", "Data-driven decisions for growth", "Clinic Owner", "Yes"],
]

for i, m in enumerate(marketing):
    r = i + 2
    for j, v in enumerate(m):
        ws5.cell(row=r, column=j + 1, value=v)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws5, r, len(mkt_headers), fill)

mkt_widths = [20, 35, 65, 45, 20, 18]
for i, w in enumerate(mkt_widths, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w
ws5.freeze_panes = "A2"

# ═══════════════════════════════════════════
# SHEET 6: TESTING TRACKER
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("Testing Tracker")
ws6.sheet_properties.tabColor = "C62828"

test_headers = ["Test ID", "Module", "Test Scenario", "Steps to Test", "Expected Result", "Tester", "Status", "Date Tested", "Bug ID", "Notes"]
for i, h in enumerate(test_headers, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, len(test_headers))

tests = [
    ["T001", "Patients", "NRIC valid checksum accepted", "Enter S1234567D in NRIC field, tab out", "No error shown, green border", "", "Pending", "", "", ""],
    ["T002", "Patients", "NRIC invalid checksum rejected", "Enter S1234567A in NRIC field, tab out", "Red border + 'Invalid NRIC checksum' error", "", "Pending", "", "", ""],
    ["T003", "Patients", "NRIC auto-uppercase", "Type s1234567d in NRIC field", "Auto-converts to S1234567D", "", "Pending", "", "", ""],
    ["T004", "Patients", "NRIC wrong prefix rejected", "Enter X1234567A, tab out", "Error: 'NRIC must start with S, T, F, G, or M'", "", "Pending", "", "", ""],
    ["T005", "Patients", "NRIC duplicate hard block", "Enter existing NRIC, try to save", "Red warning with link to existing patient, save blocked", "", "Pending", "", "", ""],
    ["T006", "Patients", "Phone duplicate soft warning", "Enter existing phone number", "Yellow warning with 'Different person, continue' option", "", "Pending", "", "", ""],
    ["T007", "Patients", "Phone normalization (SG)", "Enter 91234567 as phone, save patient", "Stored as +6591234567 in database", "", "Pending", "", "", "Check patient detail page"],
    ["T008", "Patients", "Phone normalization (IN)", "Enter 9876543210 as phone, save patient", "Stored as +919876543210 in database", "", "Pending", "", "", ""],
    ["T009", "Patients", "Secondary ≠ Primary phone", "Enter same number in Mobile and Secondary Mobile, save", "Error: 'Secondary mobile cannot be the same as primary'", "", "Pending", "", "", ""],
    ["T010", "Patients", "Emergency ≠ Own phone", "Enter same number in Mobile and Emergency Phone, save", "Error: 'Emergency contact should be different from patient's own number'", "", "Pending", "", "", ""],
    ["T011", "Patients", "Email duplicate soft warning", "Enter existing email", "Yellow warning with override option", "", "Pending", "", "", ""],
    ["T012", "Patients", "Required fields validation", "Leave First Name, Last Name, Gender, Phone empty and save", "All 4 fields show error messages", "", "Pending", "", "", ""],
    ["T013", "Inventory", "Inventory list loads with A-Z sorting", "Go to Inventory page", "Items listed alphabetically, search works, categories filter", "", "Pending", "", "", ""],
    ["T014", "Inventory", "Create new inventory item", "Inventory > New > Fill form > Save", "Item created with name, category, unit, reorder point, supplier", "", "Pending", "", "", ""],
    ["T015", "Inventory", "Add variant to item", "Edit item > Add variant (e.g., 100ml, 200ml)", "Variants appear in item detail and billing picker", "", "Pending", "", "", ""],
    ["T016", "Inventory", "Item detail — sub-tabs work", "Click item > Check Overview, Transactions, Movement Graph tabs", "All tabs load correct data, graph shows stock history", "", "Pending", "", "", ""],
    ["T017a", "Inventory", "CSV bulk import", "Inventory > Import > Upload CSV", "Items imported, validation errors shown, duplicates caught", "", "Pending", "", "", ""],
    ["T018", "Inventory", "Stock audit with barcode scan", "Stock Audit > Scan barcode > Enter count", "Item found by barcode, discrepancy calculated vs system stock", "", "Pending", "", "", ""],
    ["T019", "Inventory", "Create purchase order", "PO > New > Select supplier, add items, save", "PO created in Draft status, items and quantities correct", "", "Pending", "", "", ""],
    ["T020a", "Inventory", "PO status workflow", "Submit PO > Receive partial > Receive remaining", "Status: Draft → Submitted → Partial → Received, stock updated", "", "Pending", "", "", ""],
    ["T021", "Inventory", "Smart PO suggestions", "PO > Check suggestions", "Items below reorder point suggested with recommended quantities", "", "Pending", "", "", ""],
    ["T022", "Inventory", "Create stock transfer", "Transfers > New > Select branches, add items, submit", "Stock deducted from source branch immediately", "", "Pending", "", "", ""],
    ["T023", "Inventory", "Receive stock transfer", "Open pending transfer at dest branch > Receive", "Stock added to destination, transfer marked Received", "", "Pending", "", "", ""],
    ["T024", "Inventory", "Cancel stock transfer", "Open submitted transfer > Cancel", "Stock restored to source branch, transfer marked Cancelled", "", "Pending", "", "", ""],
    ["T025", "Inventory", "Transfer template save & apply", "Create transfer > Save as template > New transfer > Apply template", "Template pre-fills items, quantities, branches", "", "Pending", "", "", ""],
    ["T026", "Inventory", "Transfer notification", "Submit transfer to another branch", "Bell icon shows unread notification at destination branch", "", "Pending", "", "", ""],
    ["T027", "Inventory", "Multi-branch stock comparison", "Branch Stock > Comparison view", "Side-by-side stock levels for all branches, low stock highlighted", "", "Pending", "", "", ""],
    ["T028", "Inventory", "Expiry alert triggers", "Add item with expiry < 30 days", "Appears in Alerts dashboard under expiry section", "", "Pending", "", "", ""],
    ["T029", "Inventory", "Low stock alert", "Reduce stock below reorder point", "Appears in Alerts + Doctor Portal dashboard card", "", "Pending", "", "", ""],
    ["T030", "Inventory", "Write off expired stock", "Alerts > Expiry > Write Off", "Stock quantity reduced, write-off transaction recorded", "", "Pending", "", "", ""],
    ["T031", "Inventory", "Supplier CRUD", "Suppliers page > Add/Edit/Delete supplier", "Supplier saved, appears in PO supplier dropdown", "", "Pending", "", "", ""],
    ["T032", "Inventory", "Inventory reports page", "Reports > Inventory", "Valuation, movement, expiry, low-stock reports with date filters", "", "Pending", "", "", ""],
    ["T033", "Inventory", "Transfer reports", "Reports > Transfers", "Transfer volume, branch activity with date/branch filters", "", "Pending", "", "", ""],
    ["T034", "Inventory", "Branch-aware PO", "Create PO with branch selection", "Received stock goes to selected branch only", "", "Pending", "", "", ""],
    ["T035", "Inventory", "InventoryTabs navigation", "Click through all inventory sub-pages", "Tabs highlight correctly, no duplicate tabs, consistent headers", "", "Pending", "", "", ""],
    ["T040", "Billing", "Create and pay invoice", "Billing > New > Add items > Save > Record payment", "Invoice marked as paid, payment recorded", "", "Pending", "", "", ""],
    ["T041", "Billing", "Branch filter on billing", "Billing page > Select branch from dropdown", "Only invoices for selected branch shown", "", "Pending", "", "", ""],
    ["T042", "Billing", "Credit note creation", "Open paid invoice > Issue credit note", "Credit note created, linked to original invoice", "", "Pending", "", "", ""],
    ["T043", "Admin", "Login with 2FA", "Login with credentials, enter TOTP code", "Successful login, redirected to dashboard", "", "Pending", "", "", ""],
    ["T044", "Admin", "Role-based access", "Login as doctor, try to access admin page", "Access denied / redirected to doctor portal", "", "Pending", "", "", ""],
    ["T045", "Admin", "Branch management", "Admin > Branches > Create new branch", "Branch appears in all branch dropdowns", "", "Pending", "", "", ""],
    ["T046", "Admin", "Staff invite flow", "Admin > Staff > Invite new user", "Invite email sent, user can set password via link", "", "Pending", "", "", ""],
    ["T047", "Admin", "Audit log records actions", "Perform any CRUD action, check Admin > Audit Log", "Action logged with user, timestamp, details", "", "Pending", "", "", ""],
    ["T048", "Doctor Portal", "Doctor dashboard loads", "Login as doctor, check dashboard", "Today's appointments, low stock alerts, quick actions shown", "", "Pending", "", "", ""],
    ["T049", "Doctor Portal", "Write prescription", "Open patient > Prescriptions > New", "Prescription created with items, dosage, duration", "", "Pending", "", "", ""],
    ["T050", "Communications", "Send WhatsApp message", "Patient detail > Communications > Send WhatsApp", "Message sent via WhatsApp API", "", "Pending", "", "", ""],
    ["T051", "Packages", "Purchase and use package", "Packages > Buy for patient > Book session", "Session counted, remaining sessions decremented", "", "Pending", "", "", ""],
]

for i, t in enumerate(tests):
    r = i + 2
    for j, v in enumerate(t):
        ws6.cell(row=r, column=j + 1, value=v)
    fill = LIGHT_GRAY if i % 2 == 0 else WHITE
    style_row(ws6, r, len(test_headers), fill)
    status_cell = ws6.cell(row=r, column=7)
    if t[6] == "Passed":
        status_cell.fill = PatternFill("solid", fgColor=GREEN_BG)
    elif t[6] == "Failed":
        status_cell.fill = PatternFill("solid", fgColor="FFCDD2")
    elif t[6] == "Pending":
        status_cell.fill = PatternFill("solid", fgColor=YELLOW_BG)
    status_cell.alignment = center_wrap

test_widths = [10, 14, 30, 45, 40, 14, 10, 12, 10, 30]
for i, w in enumerate(test_widths, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.auto_filter.ref = f"A1:{get_column_letter(len(test_headers))}{len(tests) + 1}"
ws6.freeze_panes = "A2"

# ═══════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════
output = "/Users/karthik/Cladue CODE1/patient-registration/docs/Ayur_Centre_Project_Tracker.xlsx"
wb.save(output)
print(f"Saved: {output}")
